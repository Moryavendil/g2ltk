# -*- coding: utf-8 -*-
# <nbformat>3.0</nbformat>
# <codecell>

# %matplotlib notebook

import os
import numpy as np
import matplotlib.pyplot as plt

from tools import set_verbose, datareading, datasaving, utility
utility.configure_mpl()


# <markdowncell>

# # MEASURE DES VARIABLES
# Automated measurements of lots of things


# <codecell>

### Datasets display
root_path = '../'
datasets = datareading.find_available_datasets(root_path)
print('Available datasets:', datareading.find_available_datasets(root_path))


# <codecell>

### Dataset selection & acquisitions display
dataset = '-'
if len(datasets) == 1:
    dataset = datasets[0]
    datareading.log_info(f'Auto-selected dataset {dataset}')
dataset_path = os.path.join(root_path, dataset)
datareading.describe_dataset(dataset_path, type='gcv', makeitshort=True)


# <codecell>

### Acquisition selection
straight = ['20eta', '20seuil', '30eta', '30seuil', '40eta', '50eta', '70eta', '100eta', '200eta']
instab = ['20down', '30down', '30max', '40seuil', '40down', '40fix', '50seuil', '50mid', '50high', '70seuil', '70mid', '100seuil', '100mid', '200seuil', '200mid']
acquisition = '70seuil' + '_gcv'
acquisition_path = os.path.join(dataset_path, acquisition)


# <codecell>

### meta-info reading
import pandas as pd

metainfo = pd.read_excel(os.path.join(dataset_path, 'datasheet.xlsx'), sheet_name='metainfo', skiprows=2)

meaningful_keys = [key for key in metainfo.keys() if 'unnamed' not in key.lower()]
valid = metainfo['acquisition_title'].astype(str) != 'nan'

acquisition_title = metainfo['acquisition_title'][valid].to_numpy()
# n_pump       = metainfo['n_pump'][valid].to_numpy()
# acquisition_frequency        = metainfo['acquisition_frequency'][valid].to_numpy()
# exposure_time      = metainfo['exposure_time'][valid].to_numpy()
# n_frames       = metainfo['n_frames'][valid].to_numpy()
# excitation_frequency      = metainfo['excitation_frequency'][valid].to_numpy()
# excitation_amplitude        = metainfo['excitation_amplitude'][valid].to_numpy()
# 
# if acquisition.replace('_gcv', '') in acquisition_title:
#     utility.log_info('METADATA FROM DATASHEET FILE')
#     i_acquisition = np.where(acquisition_title == acquisition.replace('_gcv', ''))[0][0]
#     for key in meaningful_keys:
#         utility.log_info(f'{key}: {metainfo[key][i_acquisition]}')


# <codecell>

### Parameters definition

# parameters to find the rivulet
rivfinding_params = {
    'resize_factor': 2,
    'borders_min_distance': 8,
    'max_borders_luminosity_difference': 50,
    'max_rivulet_width': 100.,
}

# portion of the video that is of interest to us
framenumbers = np.arange(datareading.get_number_of_available_frames(acquisition_path))
roi = None, None, None, None  #start_x, start_y, end_x, end_y

# framenumbers = np.arange(100)
if dataset == '20241104':
    roi = 250, None, 1150, None
    if acquisition == '20down' + '_gcv':
        roi = [roi[0] + 300, *roi[1:]]



# <codecell>

# Data fetching
length, height, width = datareading.get_geometry(acquisition_path, framenumbers = framenumbers, subregion=roi)

t = datareading.get_t_frames(acquisition_path, framenumbers=framenumbers)
x = datareading.get_x_px(acquisition_path, framenumbers = framenumbers, subregion=roi, resize_factor=rivfinding_params['resize_factor'])

# conversion factor
px_per_mm = 33.6
px_per_um = px_per_mm * 1e3
acquisition_frequency = datareading.get_acquisition_frequency(acquisition_path)

z_raw = datasaving.fetch_or_generate_data('bol', dataset, acquisition, framenumbers=framenumbers, roi=roi, **rivfinding_params)
w_raw = utility.w_from_borders(datasaving.fetch_or_generate_data('borders', dataset, acquisition, framenumbers=framenumbers, roi=roi, **rivfinding_params))


# <markdowncell>

# ## data cleaning
# 
# We do various cleaning steps to obtain a cleaner $Z$ and $W$


# <codecell>

# Spatial cleaning : parabolic fit on the data

# get temporal mean
z_xprofile = np.mean(z_raw, axis=0)

# fit
rivs_fit_spatial = np.polyfit(x, z_xprofile, deg = 1)
print(f'Position spatial drift estimation: {round(np.poly1d(rivs_fit_spatial)(x).max() - np.poly1d(rivs_fit_spatial)(x).min(),2)} px')

# correct
do_xcorrect= True
z_x_treated = z_raw
w_x_treated = w_raw
if do_xcorrect:
    z_x_treated = z_raw - np.expand_dims(np.poly1d(rivs_fit_spatial)(x), axis=0)
else:
    print('No spatial correction made')


# plot
plt.figure(figsize=(8,3))
ax = plt.gca()
if do_xcorrect:
    ax.plot(x, z_xprofile, color='k', alpha=0.5, label='old time-averaged riv position')
    plt.plot(x, np.poly1d(rivs_fit_spatial)(x), color='r', alpha=0.5, label=f'linear fit')
ax.plot(x, z_x_treated.mean(axis=0), color='k', label='time-averaged riv position')
ax.set_xlabel('x (px)')
ax.set_ylabel('z (px)')
ax.legend()
plt.tight_layout()


# <codecell>

# Temporal cleaning : parabolic fit on the data

# get spatial mean
z_tprofile = np.mean(z_x_treated, axis=1)

# fit
rivs_fit_temporal = np.polyfit(t, z_tprofile, deg = 2)
print(f'Position temporal drift estimation: {round(np.poly1d(rivs_fit_temporal)(t).max() - np.poly1d(rivs_fit_temporal)(t).min(),2)} px')

# correct
do_tcorrect= False
z_xt_treated = z_x_treated
w_xt_treated = w_x_treated
if do_tcorrect:
    z_xt_treated = z_x_treated - np.expand_dims(np.poly1d(rivs_fit_temporal)(t), axis=1)
else:
    print('No temporal correction made')


# plot
plt.figure(figsize=(8,3))
ax = plt.gca()
if do_tcorrect:
    ax.plot(t, z_tprofile, color='k', alpha=0.5, label='old space-averaged riv position')
    plt.plot(t, np.poly1d(rivs_fit_temporal)(t), color='r', alpha=0.5, label=f'paraboloidal fit')
ax.plot(t, z_xt_treated.mean(axis=1), color='k', label='space-averaged riv position')
ax.set_xlabel('t (s)')
ax.set_ylabel('z (px)')
ax.legend()
plt.tight_layout()


# <codecell>

from scipy.ndimage import gaussian_filter
blur_t_frame = 3 # blur in time (frame).
sigma_t = blur_t_frame
blur_x_px = 3 # blur in space (px).
sigma_x = blur_x_px

apply_gaussianfiler = False

z_filtered = gaussian_filter(z_xt_treated, sigma=(sigma_t, sigma_x))
w_filtered = gaussian_filter(w_xt_treated, sigma=(sigma_t, sigma_x))

fig, axes = plt.subplots(2, 2, sharex=True, sharey=True, squeeze=False)
realplot_kw = {'origin': 'upper', 'interpolation': 'nearest', 'aspect': 'auto'}

ax = axes[0, 0]
ax.set_title('Z (normal)')
imz = ax.imshow(z_xt_treated, extent=utility.correct_extent(x, t), cmap='viridis', **realplot_kw)
ax.set_xlabel('$x$ [px]')
ax.set_ylabel('$t$ [frame]')
plt.colorbar(imz, ax=ax, label='$z$ [px]')

ax = axes[0, 1]
ax.set_title('W (normal)')
imw = ax.imshow(w_xt_treated, extent=utility.correct_extent(x, t), cmap='viridis', **realplot_kw)
ax.set_xlabel('$x$ [px]')
ax.set_ylabel('$t$ [frame]')
plt.colorbar(imz, ax=ax, label='$w$ [px]')

ax = axes[1, 0]
ax.set_title('Z (smoothed)')
imz = ax.imshow(z_filtered, extent=utility.correct_extent(x, t), cmap='viridis', **realplot_kw)
ax.set_xlabel('$x$ [px]')
ax.set_ylabel('$t$ [frame]')
plt.colorbar(imz, ax=ax, label='$z$ [px]')

ax = axes[1, 1]
ax.set_title('W (smoothed)')
imw = ax.imshow(w_filtered, extent=utility.correct_extent(x, t), cmap='viridis', **realplot_kw)
ax.set_xlabel('$x$ [px]')
ax.set_ylabel('$t$ [frame]')
plt.colorbar(imz, ax=ax, label='$w$ [px]')

plt.tight_layout()


# <codecell>

if apply_gaussianfiler:
    Z = z_filtered.copy()
    W = w_filtered.copy()
else:
    Z = z_xt_treated.copy()
    W = w_xt_treated.copy()


# <codecell>

fig, axes = plt.subplots(1, 2, sharex=True, sharey=True)
realplot_kw = {'origin': 'upper', 'interpolation': 'nearest', 'aspect': 'auto', 'extent': utility.correct_extent(x, t)}

ax = axes[0]
imz = ax.imshow(Z, cmap='viridis', **realplot_kw)
ax.set_xlabel('$x$ [px]')
ax.set_ylabel('$t$ [frame]')
plt.colorbar(imz, ax=ax, label='$z$ [px]')

ax = axes[1]
imw = ax.imshow(W, cmap='viridis', **realplot_kw)
ax.set_xlabel('$x$ [px]')
ax.set_ylabel('$t$ [frame]')
plt.colorbar(imw, ax=ax, label='$w$ [px]')

plt.tight_layout()


# <markdowncell>

# ## FFT
# We look at the time-space power spectra of $Z$ and $W$


# <codecell>

window = 'hann'
peak_depth_dB = 100


# <codecell>

f = utility.dual(t)
k = utility.rdual(x)

Z_psd = utility.psd2d(Z, t, x, window=window) # power spectral density
W_psd = utility.psd2d(W, t, x, window=window) # power spectral density


# <markdowncell>

# ### Heuristic finding of $f$ and $k$
# We want to automatically find the temporal and spatial frequencies


# <codecell>

### HERE WE FIND THE EXCITATION FREQUENCY BY CONSIDERING THE SPACE-AVERAGED VERSION OF Z
zmeanx = np.mean(Z, axis=1)
freq = utility.rdual(t)

# Compute the power spectral density
zmeanx_psd = utility.psd1d(zmeanx, t, window=window)  # power spectral density

# find the main peak
f0_guess = utility.find_global_max(freq, zmeanx_psd)
print(f'f_0 (guessed): {round(f0_guess, 3)} frames-1 ')

peak_edges = utility.peak_contour1d(f0_guess, zmeanx_psd, peak_depth_dB=peak_depth_dB, x=freq)
peak_indexes = utility.peak_vicinity1d(f0_guess, zmeanx_psd, peak_depth_dB=peak_depth_dB, x=freq)
z_peakpower = utility.power_near_peak1d(f0_guess, zmeanx_psd, peak_depth_dB=peak_depth_dB, x=freq)

z_peakamplitude = np.sqrt(z_peakpower) * np.sqrt(2)
print(f'z_0: {round(z_peakamplitude, 3)} px (filtering PSD of <Z>_x)')



# <codecell>

from scipy.signal import hilbert

# analytic_signal = hilbert(zmeanx) # this is brutal (but sometimes works well. it avoid the phase bad definition if at the end we are not on a min / max
analytic_signal = hilbert(np.concatenate((zmeanx, zmeanx[::-1])))[:len(zmeanx)] # We use a small symmetrization trick here bcz hilbert thinks everything in life has to be periodic smh
h = np.abs(analytic_signal)

fig = plt.figure()

# real signal 
ax = fig.add_subplot(3, 1, 1)
ax.plot(t, zmeanx, color='k')
ax.axhline(z_peakamplitude, lw=2, color='b', linestyle='--', label='$z_0$')
ax.axhline(-z_peakamplitude, lw=2, color='b', linestyle='--')
ax.plot(t, h, color='m', alpha=.5, label='Analytic signal amplitude (hilbert)')
ax.plot(t, -h, color='m', alpha=.5)
ax.set_xlabel(r'time $t$ [frame]')
ax.set_ylabel(r'<z>$_x$ [px]')
ax.legend()

# log plot, global
ax = fig.add_subplot(3, 2, 3)
ax.plot(freq, zmeanx_psd, color='k', lw=1, alpha=.8)
ax.fill_between(freq[peak_indexes], zmeanx_psd[peak_indexes], color='blue', alpha=.5)
ax.set_yscale('log')
ax.set_ylim(utility.attenuate_power(zmeanx_psd.max(), 200), zmeanx_psd.max()*2)
ax.set_xlim(0, freq.max())
ax.set_xlabel(r'frequency $f$ [frame$^{-1}$]')
ax.set_ylabel(r'[px$^2$ / frame$^{-1}$]')

# log plot, zoomed near peak
ax = fig.add_subplot(3, 2, 4)
ax.plot(freq, zmeanx_psd, color='k', lw=1, alpha=.8)
ax.axvline(f0_guess, color='blue', linestyle='--', label='Forcing freq')
# ax.axvline(peakedges[0], lw=2, color='blue', linestyle='--')
# ax.axvline(peakedges[1], lw=2, color='blue', linestyle='--')
ax.plot(freq[peak_indexes], zmeanx_psd[peak_indexes], color='k', lw=3)
ax.fill_between(freq[peak_indexes], zmeanx_psd[peak_indexes], color='blue', alpha=.5)
ax.set_yscale('log')
ax.set_ylim(utility.attenuate_power(zmeanx_psd[peak_indexes].max(), peak_depth_dB+100), zmeanx_psd[peak_indexes].max()*2)
ax.set_xlim(f0_guess - 4 * (peak_edges[1] - peak_edges[0]), f0_guess + 4 * (peak_edges[1] - peak_edges[0]))
ax.set_xlabel(r'frequency $f$ [frame$^{-1}$]')
ax.set_ylabel(r'[px$^2$ / frame$^{-1}$]')
ax.legend()

# linear plot, gloabal
ax = fig.add_subplot(3, 2, 5)
ax.plot(freq, zmeanx_psd, color='k', lw=1, alpha=.8)
ax.fill_between(freq[peak_indexes], zmeanx_psd[peak_indexes], color='blue', alpha=.5)
ax.set_ylim(0, zmeanx_psd.max()*1.15)
ax.set_xlim(0, freq.max())
ax.set_xlabel(r'frequency $f$ [frame$^{-1}$]')
ax.set_ylabel(r'[px$^2$ / frame$^{-1}$]')

# linear plot, zoomed near peak
ax = fig.add_subplot(3, 2, 6)
ax.plot(freq, zmeanx_psd, color='k', lw=1, alpha=.8)
ax.axvline(f0_guess, color='blue', linestyle='--')
# ax.axvline(peakedges[0], lw=2, color='blue', linestyle='--')
# ax.axvline(peakedges[1], lw=2, color='blue', linestyle='--')
ax.plot(freq[peak_indexes], zmeanx_psd[peak_indexes], color='k', lw=3)
ax.fill_between(freq[peak_indexes], zmeanx_psd[peak_indexes], color='blue', alpha=.5)
ax.set_ylim(0, zmeanx_psd[peak_indexes].max()*1.15)
ax.set_xlim(f0_guess - 4 * (peak_edges[1] - peak_edges[0]), f0_guess + 4 * (peak_edges[1] - peak_edges[0]))
ax.set_xlabel(r'frequency $f$ [frame$^{-1}$]')
ax.set_ylabel(r'[px$^2$ / frame$^{-1}$]')


# <codecell>

### HERE WE FIND THE WAVENUMBER BY LOOKING AT THE AVERAGED SPECTRUM OF Z

### HERE WE FIND THE EXCITATION FREQUENCY BY CONSIDERING THE SPACE-AVERAGED VERSION OF Z
zpsd_meanf = np.mean(Z_psd, axis=0)

from scipy.signal import find_peaks

mink0 = k[2] # exclude the points too close to the zero frequency

peaks_coarse = k[find_peaks(zpsd_meanf)[0]]
firstpeak_coarse = peaks_coarse[peaks_coarse > mink0].min()

peaks_fine = utility.find_extrema(k, np.log10(zpsd_meanf), 'max')
firstpeak_fine = peaks_fine[peaks_fine > mink0].min()

k0_guess = firstpeak_fine
print(f'k_0 (guessed): {round(k0_guess, 3)} px-1 ')


# <codecell>

fig, axes = plt.subplots(2, 2, sharex='col', sharey='row')

ax = axes[0, 0]
ax.plot(k, zpsd_meanf, color='k')
ax.axvline(firstpeak_coarse, lw=1, color='b', linestyle='--', alpha=0.5, label='$k$ (coarse estimation)')
ax.axvline(firstpeak_fine, lw=1, color='b', linestyle='-', alpha=1., label='$k$ (fine estimation)')
ax.set_xlim(0, k.max())
# ax.set_xlabel(r'$k$ [px$^{-1}$]')
ax.set_ylabel(r'$f$-averaged PSD of Z [px^2/frame$^{-1}$]')
ax.legend()

ax = axes[0, 1]
ax.plot(k, zpsd_meanf, color='k')
ax.axvline(firstpeak_coarse, lw=1, color='b', linestyle='--', alpha=0.5, label='$k$ (coarse estimation)')
ax.axvline(firstpeak_fine, lw=1, color='b', linestyle='-', alpha=1., label='$k$ (fine estimation)')
ax.set_xlim(0, 3.5*max(firstpeak_coarse, firstpeak_fine))
# ax.set_xlabel(r'$k$ [px$^{-1}$]')
# ax.set_ylabel(r'$f$-averaged PSD of Z [px^2/frame$^{-1}$]')
ax.legend()

ax = axes[1, 0]
ax.plot(k, zpsd_meanf, color='k')
ax.axvline(firstpeak_coarse, lw=1, color='b', linestyle='--', alpha=0.5, label='$k$ (coarse estimation)')
ax.axvline(firstpeak_fine, lw=1, color='b', linestyle='-', alpha=1., label='$k$ (fine estimation)')
ax.set_xlim(0, k.max())
ax.set_yscale('log')
ax.set_xlabel(r'$k$ [px$^{-1}$]')
ax.set_ylabel(r'$f$-averaged PSD of Z [px^2/frame$^{-1}$]')
ax.legend()

ax = axes[1, 1]
ax.plot(k, zpsd_meanf, color='k')
ax.axvline(firstpeak_coarse, lw=1, color='b', linestyle='--', alpha=0.5, label='$k$ (coarse estimation)')
ax.axvline(firstpeak_fine, lw=1, color='b', linestyle='-', alpha=1., label='$k$ (fine estimation)')
ax.set_xlim(0, 3.5*max(firstpeak_coarse, firstpeak_fine))
ax.set_yscale('log')
ax.set_xlabel(r'$k$ [px$^{-1}$]')
# ax.set_ylabel(r'$f$-averaged PSD of Z [px^2/frame$^{-1}$]')
ax.legend()



# <codecell>

### HERE WE FIND VDRIFT

### HERE WE FIND THE EXCITATION FREQUENCY BY CONSIDERING THE SPACE-AVERAGED VERSION OF Z
k_nearest = np.argmin((k-k0_guess)**2)
k_indexes = [k_nearest + i for i in [-1, 0, 1]]

zpsd_k0 = np.mean(Z_psd[:, k_indexes], axis=1)

xlim = [-f0_guess, f0_guess]
in_xlim = (f > xlim[0]) * (f < xlim[1])

fdrift_guess = utility.find_global_peak(f[in_xlim], np.log10(zpsd_k0[in_xlim]), 'max')

vdrift = -fdrift_guess / k0_guess

print(f'fdrift (guessed): {round(fdrift_guess, 3)} frame-1 ')
print(f'vdrift (guessed): {round(k0_guess, 3)} px/frame ')


# <codecell>

fig = plt.figure()

ax = fig.add_subplot(211)
ax.plot(f, zpsd_k0, color='k')
for i_f0 in range(int(f.min()/f0_guess) - 1, int(f.max()/f0_guess) + 1 + 1):
    ax.axvline(fdrift_guess + i_f0*f0_guess, lw=1, color='b', linestyle='-', alpha=1.)
ax.set_yscale('log')
ax.set_xlim(-5*f0_guess, 5*f0_guess)
ax.set_xlabel(r'$f$ [frame$^{-1}$]')
# ax.set_ylabel(r'$f$-averaged PSD of Z [px^2/frame$^{-1}$]')
# ax.legend()

ax = fig.add_subplot(223)
ax.plot(f, zpsd_k0, color='k')
# ax.axvline(firstpeak_coarse, lw=1, color='b', linestyle='--', alpha=0.5, label='$k$ (coarse estimation)')
ax.axvline(fdrift_guess, lw=1, color='b', linestyle='-', alpha=1., label='$fdrift$ (fine estimation)')
ax.set_xlim(xlim)
ax.set_xlabel(r'$f$ [frame$^{-1}$]')
# ax.set_ylabel(r'$f$-averaged PSD of Z [px^2/frame$^{-1}$]')
# ax.legend()

ax = fig.add_subplot(224)
ax.plot(f, zpsd_k0, color='k')
# ax.axvline(firstpeak_coarse, lw=1, color='b', linestyle='--', alpha=0.5, label='$k$ (coarse estimation)')
ax.axvline(fdrift_guess, lw=1, color='b', linestyle='-', alpha=1., label='$fdrift$ (fine estimation)')
ax.set_yscale('log')
ax.set_xlim(xlim)
ax.set_xlabel(r'$f$ [frame$^{-1}$]')
# ax.set_ylabel(r'$f$-averaged PSD of Z [px^2/frame$^{-1}$]')
# ax.legend()


# <codecell>

n_k = np.arange(0, 2+1)
n_f = np.arange(-2, 1+1)

mode_if, mode_ik = np.meshgrid(n_f, n_k)

mode_f = mode_ik*fdrift_guess + mode_if*f0_guess
mode_k = mode_ik*k0_guess

print(mode_ik.shape)


peak_depth_dB = 40


# <codecell>

fig, axes = plt.subplots(1, 2, sharex=True, sharey=True)
fftplot_kw = {'origin': 'lower', 'aspect': 'auto', 'interpolation': 'nearest', 'extent': utility.correct_extent(k, f, origin='lower')}
# ticks for fft
def get_cbticks(vmax, range_db):
    step_db = 5 * int(range_db / 25)
    z_ticks_db = np.arange(0, range_db, step_db)
    cbticks = [ utility.attenuate_power(vmax, att_db) for att_db in z_ticks_db]
    cbticklabels = ['ref' if att_db == 0 else f'-{att_db} dB' for att_db in z_ticks_db]
    return cbticks, cbticklabels

range_db = 160

ax = axes[0]
im_zpw = ax.imshow(Z_psd, cmap='viridis', norm='log', vmax=Z_psd.max(), vmin= utility.attenuate_power(Z_psd.max(), range_db), **fftplot_kw)
for i_k in range(len(n_k)):
    for i_f in range(len(n_f)):
        ax.scatter(mode_k[i_k, i_f], mode_f[i_k, i_f], s=30, fc='r', ec='k', lw=2)
        ax.text(mode_k[i_k, i_f], mode_f[i_k, i_f], f'({mode_ik[i_k, i_f]}, {mode_if[i_k, i_f]})', color='w', alpha=.8)
ax.set_xlabel(r'$k$ [px$^{-1}$]')
ax.set_ylabel(r'$f$ [frame$^{-1}$]')
cb = plt.colorbar(im_zpw, ax=ax, label='amp²')
cbticks, cbticklabels = get_cbticks(Z_psd.max(), range_db)
cb.ax.set_yticks(cbticks) ; cb.ax.set_yticklabels(cbticklabels)
cb.ax.minorticks_off()

ax = axes[1]
im_wpw = ax.imshow(W_psd, cmap='viridis', norm='log', vmax=W_psd.max(), vmin= utility.attenuate_power(W_psd.max(), range_db), **fftplot_kw)
for i_k in range(len(n_k)):
    for i_f in range(len(n_f)):
        ax.scatter(mode_k[i_k, i_f], mode_f[i_k, i_f], s=30, fc='r', ec='k', lw=2)
        ax.text(mode_k[i_k, i_f], mode_f[i_k, i_f], f'({mode_ik[i_k, i_f]}, {mode_if[i_k, i_f]})', color='w', alpha=.8)
ax.set_xlabel(r'$k$ [px$^{-1}$]')
ax.set_ylabel(r'$f$ [frame$^{-1}$]')
cb = plt.colorbar(im_wpw, ax=ax, label='amp²')
cbticks, cbticklabels = get_cbticks(W_psd.max(), range_db)
cb.ax.set_yticks(cbticks) ; cb.ax.set_yticklabels(cbticklabels)
cb.ax.minorticks_off()

ax.set_xlim(-utility.step(k)/2, 3.5*k0_guess)
ax.set_ylim(-3.5*f0_guess, 3.5*f0_guess)

# plt.tight_layout()


# <markdowncell>

# ### Amplitude of each mode
# We find the amplitude of each mode 


# <codecell>

nk = 1
nf = -2

i_k = np.where(n_k == nk)[0][0]
i_f = np.where(n_f == nf)[0][0]
# print(i_k, i_f)

# z0
kcentre = mode_k[i_k, i_f]
fcentre = mode_f[i_k, i_f]
label = f'  ({mode_ik[i_k, i_f]}, {mode_if[i_k, i_f]})'

from tools import set_verbose

options = {'peak_max_area': 100, 'peak_min_circularity': .4}

z_peakcontours = utility.peak_contour2d(kcentre, fcentre, Z_psd, peak_depth_dB, x=k, y=f, **options)
z_peakmask = utility.peak_vicinity2d(kcentre, fcentre, Z_psd, peak_depth_dB, x=k, y=f, peak_contours=z_peakcontours, **options)
z_peakpower = utility.power_near_peak2d(kcentre, fcentre, Z_psd, peak_depth_dB, x=k, y=f, peak_vicinity=z_peakmask, **options)

w_peakcontours = utility.peak_contour2d(kcentre, fcentre, W_psd, peak_depth_dB, x=k, y=f, **options)
w_peakmask = utility.peak_vicinity2d(kcentre, fcentre, W_psd, peak_depth_dB, x=k, y=f, peak_contours=w_peakcontours, **options)
w_peakpower = utility.power_near_peak2d(kcentre, fcentre, W_psd, peak_depth_dB, x=k, y=f, peak_vicinity=w_peakmask, **options)

z_peakamplitude = np.sqrt(z_peakpower) * np.sqrt(2)
w_peakamplitude = np.sqrt(w_peakpower) * np.sqrt(2)
utility.log_info(f'z_({mode_ik[i_k, i_f]},{mode_if[i_k, i_f]}) = {round(z_peakamplitude, 3)} px (filtering PSD of Z)')
utility.log_info(f'w_({mode_ik[i_k, i_f]},{mode_if[i_k, i_f]}) = {round(w_peakamplitude, 3)} px (filtering PSD of W)')


# <codecell>

fig, axes = plt.subplots(2, 2, sharex=True, sharey=True)
fftplot_kw = {'origin': 'lower', 'aspect': 'auto', 'interpolation': 'nearest', 'extent': utility.correct_extent(k, f, origin='lower')}

range_db = peak_depth_dB + 60

zref = Z_psd.max()
zref = Z_psd[np.argmin((f-fcentre)**2)][np.argmin((k-kcentre)**2)]

ax = axes[0, 0]
ax.set_title('Power spectrum of $z$')
im_zpw = ax.imshow(Z_psd, cmap='viridis', norm='log', vmax=zref, vmin= utility.attenuate_power(zref, range_db), **fftplot_kw)
for contour in z_peakcontours:
    utility.draw_multipolygon_edge(ax, contour, xmin=-utility.step(k)/2, lw=3, color='w')
    utility.draw_multipolygon_edge(ax, contour, xmin=-utility.step(k)/2, lw=1, color='k')
ax.scatter(kcentre, fcentre, s=30, fc='r', ec='k', lw=2)
ax.text(kcentre, fcentre, label, color='w', alpha=.8)
# ax.set_xlabel(r'$k$ [px$^{-1}$]')
ax.set_ylabel(r'$f$ [frame$^{-1}$]')
cb = plt.colorbar(im_zpw, ax=ax, label='px$^2$ / frames$^{-1}$')
cbticks, cbticklabels = get_cbticks(zref, range_db)
cb.ax.set_yticks(cbticks) ; cb.ax.set_yticklabels(cbticklabels)
cb.ax.minorticks_off()

ax = axes[0, 1]
ax.set_title('Peak region(z)')
im_wpw = ax.imshow(z_peakmask, cmap='binary_r', vmax=1.5, vmin=0., **fftplot_kw)
for contour in z_peakcontours:
    utility.draw_multipolygon_edge(ax, contour, xmin=-utility.step(k)/2, lw=3, color='w')
    utility.draw_multipolygon_edge(ax, contour, xmin=-utility.step(k)/2, lw=1, color='k')
# ax.set_xlabel(r'$k$ [px$^{-1}$]')
# ax.set_ylabel(r'$f$ [frame$^{-1}$]')

wref = W_psd.max()
wref = W_psd[np.argmin((f-fcentre)**2)][np.argmin((k-kcentre)**2)]

ax = axes[1, 0]
ax.set_title('Power spectrum of $w$')
im_wpw = ax.imshow(W_psd, cmap='viridis', norm='log', vmax=wref, vmin= utility.attenuate_power(wref, range_db), **fftplot_kw)
for contour in w_peakcontours:
    utility.draw_multipolygon_edge(ax, contour, xmin=-utility.step(k)/2, lw=3, color='w')
    utility.draw_multipolygon_edge(ax, contour, xmin=-utility.step(k)/2, lw=1, color='k')
ax.scatter(kcentre, fcentre, s=30, fc='r', ec='k', lw=2)
ax.text(kcentre, fcentre, label, color='w', alpha=.8)
ax.set_xlabel(r'$k$ [px$^{-1}$]')
ax.set_ylabel(r'$f$ [frame$^{-1}$]')
cb = plt.colorbar(im_wpw, ax=ax, label='px$^2$ / frames$^{-1}$')
cbticks, cbticklabels = get_cbticks(wref, range_db)
cb.ax.set_yticks(cbticks) ; cb.ax.set_yticklabels(cbticklabels)
cb.ax.minorticks_off()

ax = axes[1, 1]
ax.set_title('Peak region (w)')
im_wpw = ax.imshow(w_peakmask, cmap='binary_r', vmax=1.5, vmin=0., **fftplot_kw)
for contour in w_peakcontours:
    utility.draw_multipolygon_edge(ax, contour, xmin=-utility.step(k)/2, lw=3, color='w')
    utility.draw_multipolygon_edge(ax, contour, xmin=-utility.step(k)/2, lw=1, color='k')
ax.set_xlabel(r'$k$ [px$^{-1}$]')
# ax.set_ylabel(r'$f$ [frame$^{-1}$]')


ax.set_xlim(0, 1/20)
ax.set_ylim(-.1, .1)

# plt.tight_layout()


# <codecell>

import pandas as pd
### AUTODATA WORKSHEET CREATION
sheet_name = 'autodata_3x4'

workbookpath = os.path.join(dataset_path, 'datasheet.xlsx')

from openpyxl import load_workbook
 
### Check if sheet already exists
wb = load_workbook(workbookpath, read_only=True)   # open an Excel file and return a workbook
must_create_sheet = sheet_name not in wb.sheetnames
wb.close()

### Create sheet if we must.
if must_create_sheet:
    dataauto = {'acquisition_title': acquisition_title}
    for i_k, nk in enumerate(n_k):
        for i_f, nf in enumerate(n_f):
            kcentre = mode_k[i_k, i_f]
            fcentre = mode_f[i_k, i_f]
            modelabel = f'({mode_ik[i_k, i_f]},{mode_if[i_k, i_f]})'
            dataauto[f'{modelabel}_Z_amp'] = np.full(len(acquisition_title), np.nan)
            dataauto[f'{modelabel}_W_amp'] = np.full(len(acquisition_title), np.nan)
            dataauto[f'{modelabel}_freqt'] = np.full(len(acquisition_title), np.nan)
            dataauto[f'{modelabel}_freqx'] = np.full(len(acquisition_title), np.nan)
    
    df_autodata = pd.DataFrame(dataauto)
    
    with pd.ExcelWriter(workbookpath, engine='openpyxl', mode='a', if_sheet_exists='error') as writer:
        df_autodata.to_excel(writer, sheet_name=sheet_name, index=False)
        
### load sheet
df_autodata = pd.read_excel(workbookpath, sheet_name=sheet_name)

### find the right index corresponding to the current acquisition
if acquisition.replace('_gcv', '') in acquisition_title:
    i_acquisition = np.where(acquisition_title == acquisition.replace('_gcv', ''))[0][0]
else:
    raise('wtf the video is not in the meta data file ?')


# <codecell>


for i_k, nk in enumerate(n_k):
    for i_f, nf in enumerate(n_f):
        kcentre = mode_k[i_k, i_f]
        fcentre = mode_f[i_k, i_f]
        modelabel = f'({mode_ik[i_k, i_f]},{mode_if[i_k, i_f]})'
        
        i_current = i_k*len(n_f) + i_f
        i_tot = len(n_k) * len(n_f)
        pc = round(i_current/i_tot*100, 1)
        utility.log_info(f'Treating mode ({mode_ik[i_k, i_f]},{mode_if[i_k, i_f]}) ({i_current}/{i_tot} - {pc}% done)')

        #
        fcentre = mode_f[i_k, i_f]
        kcentre = mode_k[i_k, i_f]

        df_autodata.loc[i_acquisition, f'{modelabel}_freqt'] = 12

        df_autodata.loc[i_acquisition, f'{modelabel}_freqt'] = fcentre * acquisition_frequency
        df_autodata.loc[i_acquisition, f'{modelabel}_freqx'] = kcentre * px_per_mm
        
        options = {'peak_max_area': 100, 'peak_min_circularity': .4}

        z_peakamplitude = 0
        w_peakamplitude = 0
        if not(kcentre == 0 and fcentre == 0):
        
            z_peakcontours = utility.peak_contour2d(kcentre, fcentre, Z_psd, peak_depth_dB, x=k, y=f, **options)
            z_peakmask = utility.peak_vicinity2d(kcentre, fcentre, Z_psd, peak_depth_dB, x=k, y=f, peak_contours=z_peakcontours, **options)
            z_peakpower = utility.power_near_peak2d(kcentre, fcentre, Z_psd, peak_depth_dB, x=k, y=f, peak_vicinity=z_peakmask, **options)
            
            w_peakcontours = utility.peak_contour2d(kcentre, fcentre, W_psd, peak_depth_dB, x=k, y=f, **options)
            w_peakmask = utility.peak_vicinity2d(kcentre, fcentre, W_psd, peak_depth_dB, x=k, y=f, peak_contours=w_peakcontours, **options)
            w_peakpower = utility.power_near_peak2d(kcentre, fcentre, W_psd, peak_depth_dB, x=k, y=f, peak_vicinity=w_peakmask, **options)
            
            z_peakamplitude = np.sqrt(z_peakpower) * np.sqrt(2)
            w_peakamplitude = np.sqrt(w_peakpower) * np.sqrt(2)
            
        utility.log_subinfo(f'z_({mode_ik[i_k, i_f]},{mode_if[i_k, i_f]}) = {round(z_peakamplitude, 3)} px (filtering PSD of Z)')
        utility.log_subinfo(f'w_({mode_ik[i_k, i_f]},{mode_if[i_k, i_f]}) = {round(w_peakamplitude, 3)} px (filtering PSD of W)')

        df_autodata.loc[i_acquisition, f'{modelabel}_Z_amp'] = z_peakamplitude / px_per_mm
        df_autodata.loc[i_acquisition, f'{modelabel}_W_amp'] = w_peakamplitude / px_per_mm



# <codecell>


with pd.ExcelWriter(workbookpath, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    df_autodata.to_excel(writer, sheet_name=sheet_name, index=False)


# <codecell>



