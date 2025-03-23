# -*- coding: utf-8 -*-
# <nbformat>3.0</nbformat>
# <codecell>

# %matplotlib notebook

import numpy as np
import matplotlib.pyplot as plt

from g2ltk import datareading, datasaving, utility, logging
utility.configure_mpl()


# <codecell>

### Datasets display
datareading.set_default_root_path('../')
datareading.describe_root_path()


# <codecell>

### Dataset selection & acquisitions display
dataset = '40evo'

datareading.describe_dataset(dataset=dataset, videotype='gcv', makeitshort=True)
dataset_path = datareading.generate_dataset_path(dataset)


# <codecell>

### Acquisition selection
acquisition = 'a120'
acquisition_path = datareading.generate_acquisition_path(acquisition, dataset)


# <codecell>

### Parameters definition

# conversion factor
px_per_mm = (1563/30+1507/30)/2
px_per_um = px_per_mm * 1e3

# parameters to find the rivulet
rivfinding_params = {
    'resize_factor': 2,
    'remove_median_bckgnd_zwise': True,
    'gaussian_blur_kernel_size': (1, 1), # (sigma_z, sigma_x)
    'white_tolerance': 70,
    'borders_min_distance': 5.,
    'borders_width': 6.,
    'max_rivulet_width': 150.,
}

# portion of the video that is of interest to us
framenumbers = np.arange(datareading.get_number_of_available_frames(acquisition_path))
roi = None, None, None, None  #start_x, start_y, end_x, end_y
if dataset=='40evo' and acquisition=='a350':
    framenumbers = np.arange(450)



# <codecell>

# Data fetching
datareading.describe_acquisition(dataset, acquisition, framenumbers = framenumbers, subregion=roi)

length, height, width = datareading.get_geometry(acquisition_path, framenumbers = framenumbers, subregion=roi)

t = datareading.get_t_frames(acquisition_path, framenumbers=framenumbers)
x = datareading.get_x_px(acquisition_path, framenumbers = framenumbers, subregion=roi, resize_factor=rivfinding_params['resize_factor'])

z_raw = datasaving.fetch_or_generate_data('bol', dataset, acquisition, framenumbers=framenumbers, roi=roi, **rivfinding_params)
w_raw = datasaving.fetch_or_generate_data('fwhmol', dataset, acquisition, framenumbers=framenumbers, roi=roi, **rivfinding_params)



# <codecell>

z_tmp = z_raw.copy()
w_tmp = w_raw.copy()


# <codecell>

from scipy.ndimage import gaussian_filter
blur_t_frame = 2 # blur in time (frame).
sigma_t = blur_t_frame
blur_x_px = 5 # blur in space (px).
sigma_x = blur_x_px


z_filtered = gaussian_filter(z_tmp, sigma=(sigma_t, sigma_x))
w_filtered = gaussian_filter(w_tmp, sigma=(sigma_t, sigma_x))

fig, axes = plt.subplots(2, 2, sharex=True, sharey=True, squeeze=False)
imshow_kw = {'aspect': 'auto', 'origin': 'upper'}

ax = axes[0, 0]
ax.set_title('Z (normal)')
imz = ax.imshow(z_tmp, extent=utility.correct_extent(x, t), cmap='viridis', **imshow_kw)
ax.set_xlabel('$x$ [px]')
ax.set_ylabel('$t$ [frame]')
plt.colorbar(imz, ax=ax, label='$z$ [px]')

ax = axes[0, 1]
ax.set_title('Z (smoothed)')
imz = ax.imshow(z_filtered, extent=utility.correct_extent(x, t), cmap='viridis', **imshow_kw)
ax.set_xlabel('$x$ [px]')
ax.set_ylabel('$t$ [frame]')
plt.colorbar(imz, ax=ax, label='$z$ [px]')

ax = axes[1, 0]
ax.set_title('W (normal)')
imz = ax.imshow(w_tmp, extent=utility.correct_extent(x, t), cmap='viridis', **imshow_kw)
ax.set_xlabel('$x$ [px]')
ax.set_ylabel('$t$ [frame]')
plt.colorbar(imz, ax=ax, label='$z$ [px]')

ax = axes[1, 1]
ax.set_title('W (smoothed)')
imz = ax.imshow(w_filtered, extent=utility.correct_extent(x, t), cmap='viridis', **imshow_kw)
ax.set_xlabel('$x$ [px]')
ax.set_ylabel('$t$ [frame]')
plt.colorbar(imz, ax=ax, label='$z$ [px]')

plt.tight_layout()

apply_gaussianfiler = True
if apply_gaussianfiler:
    z_tmp = z_filtered.copy()
    w_tmp = w_filtered.copy()
    utility.log_info('Gaussian filtering correction made')
else:
    utility.log_info('No gaussian filtering correction made')


# <codecell>

# Spatial cleaning : parabolic fit on the data
from scipy.signal import savgol_filter

# get temporal mean
z_xprofile = np.mean(z_tmp, axis=0)

# linear fit
rivs_fit_spatial = np.polyfit(x, z_xprofile, deg = 1)
utility.log_info(f'Position spatial drift linear estimation: {round(np.poly1d(rivs_fit_spatial)(x).max() - np.poly1d(rivs_fit_spatial)(x).min(),2)} px')

# savgol
savgol_width = len(x)//10 + (1 if len(x)%2 == 0 else 0)

"""
xcorrect can be 3 things
 * None: we do not want correction : it is biaised, or we are interested in the mean-time features
 * 'linear': we do a linear fit and remove everything from it
 * 'smoothed': we do a smoothing avec the average and remove that
 * 'total': total removal of the mean value
"""
xcorrect = 'linear'
if xcorrect is None:
    utility.log_info('No spatial correction made')
elif xcorrect == 'linear':
    utility.log_info('Linear spatial correction made')
    z_tmp = z_tmp - np.expand_dims(np.poly1d(rivs_fit_spatial)(x), axis=0)
elif xcorrect == 'smoothed':
    utility.log_info(f'Smoothed spatial correction made (savgol-{savgol_width}-2)')
    z_tmp = z_tmp - savgol_filter(z_xprofile, savgol_width, 2)
elif xcorrect == 'total':
    utility.log_info('Total spatial correction made')
    z_tmp = z_tmp - z_xprofile
else:
    utility.log_warning(f'What do you mean by xcorrect={xcorrect} ?')
    utility.log_info('No spatial correction made')


# plot
fig, axes = plt.subplots(2, 1, sharex=True, figsize=utility.figsize('double'))
ax = axes[0]
ax.plot(x, z_xprofile, color='k', alpha=0.5, label='old time-averaged riv position')
ax.plot(x, np.poly1d(rivs_fit_spatial)(x), color='r', alpha=0.5, label=f'linear fit')
ax.plot(x, savgol_filter(z_xprofile, savgol_width, 2), color='b', alpha=0.5, label=f'smooth')
ax.set_ylabel('z (px)')
ax.legend()

ax = axes[1]
ax.plot(x, z_tmp.mean(axis=0), color='k', label='New time-averaged riv position')
ax.set_xlabel('x (px)')
ax.set_ylabel('z (px)')
ax.legend()
plt.tight_layout()


# <codecell>

# Temporal cleaning : parabolic fit on the data

# get spatial mean
z_tprofile = np.mean(z_tmp, axis=1)

# fit
rivs_fit_temporal = np.polyfit(t, z_tprofile, deg = 2)

# correct
do_tcorrect= False
if do_tcorrect:
    z_tmp = z_tmp - np.expand_dims(np.poly1d(rivs_fit_temporal)(t), axis=1)
else:
    utility.log_info('No temporal correction made')

# plot
plt.figure(figsize=(8,3))
ax = plt.gca()
if do_tcorrect:
    ax.plot(t, z_tprofile, color='k', alpha=0.5, label='old space-averaged riv position')
    plt.plot(t, np.poly1d(rivs_fit_temporal)(t), color='r', alpha=0.5, label=f'paraboloidal fit')
ax.plot(t, z_tmp.mean(axis=1), color='k', label='space-averaged riv position')
ax.set_xlabel('t (s)')
ax.set_ylabel('z (px)')
ax.legend()
plt.tight_layout()


# <codecell>

Z = z_tmp.copy()
W = w_tmp.copy()

zero_pad_factor = (5,5)
window='hann'

k, f = utility.fourier.dual2d(x, t, zero_pad_factor=zero_pad_factor)
Z_pw = utility.fourier.psd2d(Z, x, t, window=window, zero_pad_factor=zero_pad_factor)
W_pw = utility.fourier.psd2d(W, x, t, window=window, zero_pad_factor=zero_pad_factor)

range_db = 100


# <codecell>

print('Wmean', W.mean() / px_per_mm)
print('Wstd', W.std() / px_per_mm)
print('mm_per_px', 1/px_per_mm)

plt.figure()
plt.hist(W.flatten() / px_per_mm, bins = np.linspace(W.min() / px_per_mm*.9, W.max() / px_per_mm*1.1, 200))


# <codecell>

fig, axes = plt.subplots(2, 2)
imshow_kw = {'origin':'upper', 
             'interpolation':'nearest', 
             'aspect':'auto'}

ax = axes[0,0]
ax.set_title('Z (normal)')
imz = ax.imshow(Z, extent=utility.correct_extent(x, t), cmap='viridis', **imshow_kw)
ax.set_xlabel('$x$ [px]')
ax.set_ylabel('$t$ [frame]')
plt.colorbar(imz, ax=ax, label='$z$ [px]')

ax = axes[0,1]
vmax, vmin = utility.log_amplitude_range(Z_pw.max(), range_db=range_db)
im_zpw = ax.imshow(Z_pw, extent=utility.correct_extent(k, f), norm='log', vmax=vmax, vmin=vmin, cmap='viridis', **imshow_kw)
ax.set_xlabel(r'$k$ [px$^{-1}$]')
ax.set_ylabel(r'$f$ [frame$^{-1}$]')
cb = plt.colorbar(im_zpw, ax=ax, label=r'$|\hat{z}|^2$ [px^2/(px-1.frame-1)]')
utility.set_ticks_log_cb(cb, vmax, range_db=range_db)

ax.set_xlim(0, 1/50)
ax.set_ylim(-1/20, 1/20)

ax = axes[1,0]
ax.set_title('W (normal)')
imz = ax.imshow(W, extent=utility.correct_extent(x, t), cmap='viridis', **imshow_kw)
ax.set_xlabel('$x$ [px]')
ax.set_ylabel('$t$ [frame]')
plt.colorbar(imz, ax=ax, label='$w$ [px]')

ax = axes[1,1]
vmax, vmin = utility.log_amplitude_range(W_pw.max(), range_db=range_db)
im_zpw = ax.imshow(W_pw, extent=utility.correct_extent(k, f), norm='log', vmax=vmax, vmin=vmin, cmap='viridis', **imshow_kw)
ax.set_xlabel(r'$k$ [px$^{-1}$]')
ax.set_ylabel(r'$f$ [frame$^{-1}$]')
cb = plt.colorbar(im_zpw, ax=ax, label=r'$|\hat{w}|^2$ [px^2/(px-1.frame-1)]')
utility.set_ticks_log_cb(cb, vmax, range_db=range_db)

ax.set_xlim(0, 1/50)
ax.set_ylim(-1/20, 1/20)

# plt.tight_layout()


# <markdowncell>

# ## ##  # F# i# n# d#  # $# z# _# 1# $#  # a# n# d#  # $# w# _# 1# $


# <codecell>

### meta-info reading
import pandas as pd

metainfo = pd.read_excel(os.path.join(dataset_path, 'datasheet.xlsx'), sheet_name='metainfo', skiprows=2)

meaningful_keys = [key for key in metainfo.keys() if 'unnamed' not in key.lower()]
valid = metainfo['acquisition_title'].astype(str) != 'nan'

acquisition_title = metainfo['acquisition_title'][valid].to_numpy()


# <codecell>

import pandas as pd
### AUTODATA WORKSHEET CREATION
sheet_name = 'autodata'

workbookpath = os.path.join(dataset_path, 'datasheet.xlsx')

from openpyxl import load_workbook
 
### Check if sheet already exists
wb = load_workbook(workbookpath, read_only=True)   # open an Excel file and return a workbook
must_create_sheet = sheet_name not in wb.sheetnames
wb.close()

### Create sheet if we must.
if must_create_sheet:
    dataauto = {'acquisition_title': acquisition_title}
    # fknu
    dataauto[f'f0'] = np.full(len(acquisition_title), np.nan)
    dataauto[f'q0'] = np.full(len(acquisition_title), np.nan)
    dataauto[f'fdrift'] = np.full(len(acquisition_title), np.nan)
    # amplitudes
    dataauto[f'z0'] = np.full(len(acquisition_title), np.nan)
    dataauto[f'z1'] = np.full(len(acquisition_title), np.nan)
    dataauto[f'w1'] = np.full(len(acquisition_title), np.nan)
    
    df_autodata = pd.DataFrame(dataauto)
    
    with pd.ExcelWriter(workbookpath, engine='openpyxl', mode='a', if_sheet_exists='error') as writer:
        df_autodata.to_excel(writer, sheet_name=sheet_name, index=False)
        
### load sheet
df_autodata = pd.read_excel(workbookpath, sheet_name=sheet_name)

### find the right index corresponding to the current acquisition
if acquisition.replace('_gcv', '') in acquisition_title:
    i_acquisition = np.where(acquisition_title == acquisition.replace('_gcv', ''))[0][0]
else:
    raise('wtf the video is not in the meta data file ?')

k0 = df_autodata['q0'][i_acquisition]
f0 = df_autodata['f0'][i_acquisition]
z0_global = df_autodata['z0'][i_acquisition]
L = 1/k0 * (1+0.1)


# <codecell>

zone_start = x.max() - L
zone_start = 0
zone_end = zone_start + L

zone = (x > zone_start) & (x < zone_end)



# <codecell>

## Estimate z0


# <codecell>

### HERE WE FIND THE EXCITATION FREQUENCY BY CONSIDERING THE SPACE-AVERAGED VERSION OF Z
window = 'hann'
peak_depth_dB = 60

# Take the space-average
zmeanx = np.mean(Z[:, np.argmin((x - zone_start)**2):np.argmin((x - zone_end)**2)+1], axis=1)

# Compute the power spectral density
freq = utility.rdual(t, zero_pad_factor=zero_pad_factor[0])
zmeanx_psd = utility.psd1d(zmeanx, t, window=window, zero_pad_factor=zero_pad_factor[0])  # power spectral density

# find the main peak
f0_guess = utility.fourier.estimatesignalfrequency(zmeanx, t, window='boxcar')
print(f'f_0 (guessed): {round(f0_guess, 3)} frames-1 ')

peak_edges = utility.peak_contour1d(f0_guess, zmeanx_psd, peak_depth_dB=peak_depth_dB, x=freq)
peak_indexes = utility.peak_vicinity1d(f0_guess, zmeanx_psd, peak_depth_dB=peak_depth_dB, x=freq)
z_peakpower = utility.power_near_peak1d(f0_guess, zmeanx_psd, peak_depth_dB=peak_depth_dB, x=freq)

z_peakamplitude = np.sqrt(z_peakpower) * np.sqrt(2)
print(f'z_0: {round(z_peakamplitude, 3)} px (filtering PSD of <Z>_x)')
z0_measure = z_peakamplitude


# <codecell>

## Estimate z1 via fit


# <codecell>

## Estimate w1 via fit

