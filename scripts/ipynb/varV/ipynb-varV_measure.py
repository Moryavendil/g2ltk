# -*- coding: utf-8 -*-
# <nbformat>3.0</nbformat>
# <codecell>

# %matplotlib notebook

import numpy as np
import matplotlib.pyplot as plt

from g2ltk import datareading, datasaving, utility, logging
utility.configure_mpl()


# <codecell>

### Datasets display
datareading.set_default_root_path('../')
datareading.describe_root_path()


# <codecell>

### Dataset selection & acquisitions display
dataset = '40evo'

datareading.describe_dataset(dataset=dataset, videotype='gcv', makeitshort=True)
dataset_path = datareading.generate_dataset_path(dataset)


# <codecell>

### Acquisition selection
acquisition = 'a140'
acquisition_path = datareading.generate_acquisition_path(acquisition, dataset)


# <codecell>

### Parameters definition

# conversion factor
px_per_mm = (1563/30+1507/30)/2
px_per_um = px_per_mm * 1e3

# parameters to find the rivulet
rivfinding_params = {
    'resize_factor': 2,
    'remove_median_bckgnd_zwise': True,
    'gaussian_blur_kernel_size': (1, 1), # (sigma_z, sigma_x)
    'white_tolerance': 70,
    'borders_min_distance': 5.,
    'borders_width': 6.,
    'max_rivulet_width': 150.,
}

# portion of the video that is of interest to us
framenumbers = np.arange(datareading.get_number_of_available_frames(acquisition_path))
roi = None, None, None, None  #start_x, start_y, end_x, end_y
if dataset=='40evo' and acquisition=='a350':
    framenumbers = np.arange(450)
if dataset=='40evo' and acquisition=='a320':
    framenumbers = framenumbers[framenumbers < 300]
    # framenumbers = framenumbers[framenumbers > 400]



# <codecell>

# Data fetching
datareading.describe_acquisition(dataset, acquisition, framenumbers = framenumbers, subregion=roi)

length, height, width = datareading.get_geometry(acquisition_path, framenumbers = framenumbers, subregion=roi)

t = datareading.get_t_frames(acquisition_path, framenumbers=framenumbers)
x = datareading.get_x_px(acquisition_path, framenumbers = framenumbers, subregion=roi, resize_factor=rivfinding_params['resize_factor'])

z_raw = datasaving.fetch_or_generate_data('bol', dataset, acquisition, framenumbers=framenumbers, roi=roi, **rivfinding_params)
w_raw = datasaving.fetch_or_generate_data('fwhmol', dataset, acquisition, framenumbers=framenumbers, roi=roi, **rivfinding_params)



# <codecell>

z_tmp = z_raw.copy()
w_tmp = w_raw.copy()


# <codecell>

# Step 1: Gaussian blur to remove high-frequency noise

from scipy.ndimage import gaussian_filter
blur_t_frame = 2 # blur in time (frame).
sigma_t = blur_t_frame
blur_x_px = 5 # blur in space (px).
sigma_x = blur_x_px


z_filtered = gaussian_filter(z_tmp, sigma=(sigma_t, sigma_x))
w_filtered = gaussian_filter(w_tmp, sigma=(sigma_t, sigma_x))

fig, axes = plt.subplots(2, 2, sharex=True, sharey=True, squeeze=False)
imshow_kw = {'aspect': 'auto', 'origin': 'upper'}

ax = axes[0, 0]
ax.set_title('Z (normal)')
imz = ax.imshow(z_tmp, extent=utility.correct_extent(x, t), cmap='viridis', **imshow_kw)
ax.set_xlabel('$x$ [px]')
ax.set_ylabel('$t$ [frame]')
plt.colorbar(imz, ax=ax, label='$z$ [px]')

ax = axes[0, 1]
ax.set_title('Z (smoothed)')
imz = ax.imshow(z_filtered, extent=utility.correct_extent(x, t), cmap='viridis', **imshow_kw)
ax.set_xlabel('$x$ [px]')
ax.set_ylabel('$t$ [frame]')
plt.colorbar(imz, ax=ax, label='$z$ [px]')

ax = axes[1, 0]
ax.set_title('W (normal)')
imz = ax.imshow(w_tmp, extent=utility.correct_extent(x, t), cmap='viridis', **imshow_kw)
ax.set_xlabel('$x$ [px]')
ax.set_ylabel('$t$ [frame]')
plt.colorbar(imz, ax=ax, label='$z$ [px]')

ax = axes[1, 1]
ax.set_title('W (smoothed)')
imz = ax.imshow(w_filtered, extent=utility.correct_extent(x, t), cmap='viridis', **imshow_kw)
ax.set_xlabel('$x$ [px]')
ax.set_ylabel('$t$ [frame]')
plt.colorbar(imz, ax=ax, label='$z$ [px]')

plt.tight_layout()

apply_gaussianfiler = True
if apply_gaussianfiler:
    z_tmp = z_filtered.copy()
    w_tmp = w_filtered.copy()
    utility.log_info('Gaussian filtering correction made')
else:
    utility.log_info('No gaussian filtering correction made')


# <codecell>

# Step 2: W correction due to Z slope

w_slopecorrected = w_tmp / np.sqrt(1 + np.gradient(z_tmp, axis=1)**2)

plt.figure()
wbins = np.linspace(0, w_tmp.max(), 101, endpoint=True)
ax = plt.gca()
ax.hist(w_tmp.flatten(), alpha=.5, bins=wbins, color='b')
ax.hist(w_slopecorrected.flatten(), alpha=.5, bins=wbins, color='k')

w_tmp = w_slopecorrected


# <codecell>

# Step 3: Spatial cleaning of z, we correct the angle of camera
from scipy.signal import savgol_filter

# get temporal mean
z_xprofile = np.mean(z_tmp, axis=0)

# linear fit
rivs_fit_spatial = np.polyfit(x, z_xprofile, deg = 1)
utility.log_info(f'Position spatial drift linear estimation: {round(np.poly1d(rivs_fit_spatial)(x).max() - np.poly1d(rivs_fit_spatial)(x).min(),2)} px')

# savgol
savgol_width = len(x)//10 + (1 if len(x)%2 == 0 else 0)

"""
xcorrect can be 3 things
 * None: we do not want correction : it is biaised, or we are interested in the mean-time features
 * 'linear': we do a linear fit and remove everything from it
 * 'smoothed': we do a smoothing avec the average and remove that
 * 'total': total removal of the mean value
"""
xcorrect = 'linear'
if xcorrect is None:
    utility.log_info('No spatial correction made')
elif xcorrect == 'linear':
    utility.log_info('Linear spatial correction made')
    z_tmp = z_tmp - np.expand_dims(np.poly1d(rivs_fit_spatial)(x), axis=0)
elif xcorrect == 'smoothed':
    utility.log_info(f'Smoothed spatial correction made (savgol-{savgol_width}-2)')
    z_tmp = z_tmp - savgol_filter(z_xprofile, savgol_width, 2)
elif xcorrect == 'total':
    utility.log_info('Total spatial correction made')
    z_tmp = z_tmp - z_xprofile
else:
    utility.log_warning(f'What do you mean by xcorrect={xcorrect} ?')
    utility.log_info('No spatial correction made')


# plot
fig, axes = plt.subplots(2, 1, sharex=True, figsize=utility.figsize('double'))
ax = axes[0]
ax.plot(x, z_xprofile, color='k', alpha=0.5, label='old time-averaged riv position')
ax.plot(x, np.poly1d(rivs_fit_spatial)(x), color='r', alpha=0.5, label=f'linear fit')
ax.plot(x, savgol_filter(z_xprofile, savgol_width, 2), color='b', alpha=0.5, label=f'smooth')
ax.set_ylabel('z (px)')
ax.legend()

ax = axes[1]
ax.plot(x, z_tmp.mean(axis=0), color='k', label='New time-averaged riv position')
ax.set_xlabel('x (px)')
ax.set_ylabel('z (px)')
ax.legend()
plt.tight_layout()


# <codecell>

# Step 4: Temporal cleaning : we can correct a drift in time

# get spatial mean
z_tprofile = np.mean(z_tmp, axis=1)

# fit
rivs_fit_temporal = np.polyfit(t, z_tprofile, deg = 2)

# correct
do_tcorrect= False
if do_tcorrect:
    z_tmp = z_tmp - np.expand_dims(np.poly1d(rivs_fit_temporal)(t), axis=1)
else:
    utility.log_info('No temporal correction made')

# plot
plt.figure(figsize=(8,3))
ax = plt.gca()
if do_tcorrect:
    ax.plot(t, z_tprofile, color='k', alpha=0.5, label='old space-averaged riv position')
    plt.plot(t, np.poly1d(rivs_fit_temporal)(t), color='r', alpha=0.5, label=f'paraboloidal fit')
ax.plot(t, z_tmp.mean(axis=1), color='k', label='space-averaged riv position')
ax.set_xlabel('t (s)')
ax.set_ylabel('z (px)')
ax.legend()
plt.tight_layout()


# <codecell>

Z = z_tmp.copy()
W = w_tmp.copy()

zero_pad_factor = (5,5)
window='hann'

k, f = utility.fourier.dual2d(x, t, zero_pad_factor=zero_pad_factor)
Z_pw = utility.fourier.psd2d(Z, x, t, window=window, zero_pad_factor=zero_pad_factor)
W_pw = utility.fourier.psd2d(W, x, t, window=window, zero_pad_factor=zero_pad_factor)

range_db = 40


# <codecell>

# fig, axes = plt.subplots(2, 2, sharex='col', sharey='col')
# imshow_kw = {'origin':'upper', 
#              'interpolation':'nearest', 
#              'aspect':'auto'}
# 
# ax = axes[0,0]
# ax.set_title('Z (normal)')
# imz = ax.imshow(Z, extent=utility.correct_extent(x, t), cmap='viridis', **imshow_kw)
# ax.set_xlabel('$x$ [px]')
# ax.set_ylabel('$t$ [frame]')
# plt.colorbar(imz, ax=ax, label='$z$ [px]')
# 
# ax = axes[0,1]
# vmax, vmin = utility.log_amplitude_range(Z_pw.max(), range_db=range_db)
# im_zpw = ax.imshow(Z_pw, extent=utility.correct_extent(k, f), norm='log', vmax=vmax, vmin=vmin, cmap='viridis', **imshow_kw)
# ax.set_xlabel(r'$k$ [px$^{-1}$]')
# ax.set_ylabel(r'$f$ [frame$^{-1}$]')
# cb = plt.colorbar(im_zpw, ax=ax, label=r'$|\hat{z}|^2$ [px^2/(px-1.frame-1)]')
# utility.set_ticks_log_cb(cb, vmax, range_db=range_db)
# 
# ax.set_xlim(0, 1/50)
# ax.set_ylim(-1/20, 1/20)
# 
# ax = axes[1,0]
# ax.set_title('W (normal)')
# imz = ax.imshow(W, extent=utility.correct_extent(x, t), cmap='viridis', **imshow_kw)
# ax.set_xlabel('$x$ [px]')
# ax.set_ylabel('$t$ [frame]')
# plt.colorbar(imz, ax=ax, label='$w$ [px]')
# 
# ax = axes[1,1]
# vmax, vmin = utility.log_amplitude_range(W_pw.max(), range_db=range_db)
# im_zpw = ax.imshow(W_pw, extent=utility.correct_extent(k, f), norm='log', vmax=vmax, vmin=vmin, cmap='viridis', **imshow_kw)
# ax.set_xlabel(r'$k$ [px$^{-1}$]')
# ax.set_ylabel(r'$f$ [frame$^{-1}$]')
# cb = plt.colorbar(im_zpw, ax=ax, label=r'$|\hat{w}|^2$ [px^2/(px-1.frame-1)]')
# utility.set_ticks_log_cb(cb, vmax, range_db=range_db)
# 
# ax.set_xlim(-1/50, 1/50)
# ax.set_ylim(-1/20, 1/20)
# 
# # plt.tight_layout()


# <codecell>

# Find all about F
# Assumption : homogeneous forcing

### HERE WE FIND THE EXCITATION FREQUENCY BY CONSIDERING THE SPACE-AVERAGED VERSION OF Z
window = 'hann'
peak_depth_dB = 60

# Take the space-average
zmeanx = np.mean(Z, axis=1)

# Compute the power spectral density
freq = utility.rdual(t, zero_pad_factor=zero_pad_factor[0])
zmeanx_psd = utility.psd1d(zmeanx, t, window=window, zero_pad_factor=zero_pad_factor[0])  # power spectral density

# find the main peak
f0_guess = utility.fourier.estimatesignalfrequency(zmeanx, t, window='boxcar')
print(f'f_0 (guessed): {round(f0_guess, 3)} frames-1 ')

peak_edges = utility.peak_contour1d(f0_guess, zmeanx_psd, peak_depth_dB=peak_depth_dB, x=freq)
peak_indexes = utility.peak_vicinity1d(f0_guess, zmeanx_psd, peak_depth_dB=peak_depth_dB, x=freq)
z_peakpower = utility.power_near_peak1d(f0_guess, zmeanx_psd, peak_depth_dB=peak_depth_dB, x=freq)

z_peakamplitude = np.sqrt(z_peakpower) * np.sqrt(2)
print(f'z_0: {round(z_peakamplitude, 3)} px (filtering PSD of <Z>_x)')
z0_measure = z_peakamplitude


# <codecell>

from scipy.signal import hilbert

# analytic_signal = hilbert(zmeanx) # this is brutal (but sometimes works well. it avoid the phase bad definition if at the end we are not on a min / max
analytic_signal = hilbert(np.concatenate((zmeanx, zmeanx[::-1])))[:len(zmeanx)] # We use a small symmetrization trick here bcz hilbert thinks everything in life has to be periodic smh
h = np.abs(analytic_signal)

fig = plt.figure()

# real signal 
ax = fig.add_subplot(3, 1, 1)
ax.plot(t, zmeanx, color='k')
ax.axhline(z_peakamplitude, lw=2, color='b', linestyle='--', label='$z_0$')
ax.axhline(-z_peakamplitude, lw=2, color='b', linestyle='--')
# ax.plot(t, h, color='m', alpha=.5, label='Analytic signal amplitude (hilbert)')
# ax.plot(t, -h, color='m', alpha=.5)
ax.set_xlabel(r'time $t$ [frame]')
ax.set_ylabel(r'<z>$_x$ [px]')
ax.legend()

# log plot, global
ax = fig.add_subplot(3, 2, 3)
ax.plot(freq, zmeanx_psd, color='k', lw=1, alpha=.8)
ax.fill_between(freq[peak_indexes], zmeanx_psd[peak_indexes], color='blue', alpha=.5)
ax.set_yscale('log')
ax.set_ylim(utility.attenuate_power(zmeanx_psd.max(), 200), zmeanx_psd.max()*2)
ax.set_xlim(0, freq.max())
ax.set_xlabel(r'frequency $f$ [frame$^{-1}$]')
ax.set_ylabel(r'[px$^2$ / frame$^{-1}$]')

# log plot, zoomed near peak
ax = fig.add_subplot(3, 2, 4)
ax.plot(freq, zmeanx_psd, color='k', lw=1, alpha=.8)
ax.axvline(f0_guess, color='blue', linestyle='--', label='Forcing freq')
# ax.axvline(peakedges[0], lw=2, color='blue', linestyle='--')
# ax.axvline(peakedges[1], lw=2, color='blue', linestyle='--')
ax.plot(freq[peak_indexes], zmeanx_psd[peak_indexes], color='k', lw=3)
ax.fill_between(freq[peak_indexes], zmeanx_psd[peak_indexes], color='blue', alpha=.5)
ax.set_yscale('log')
ax.set_ylim(utility.attenuate_power(zmeanx_psd[peak_indexes].max(), range_db+40), zmeanx_psd[peak_indexes].max()*2)
ax.set_xlim(max(0, f0_guess - 4 * (peak_edges[1] - peak_edges[0])), f0_guess + 4 * (peak_edges[1] - peak_edges[0]))
ax.set_xlabel(r'frequency $f$ [frame$^{-1}$]')
ax.set_ylabel(r'[px$^2$ / frame$^{-1}$]')
ax.legend()

# linear plot, gloabal
ax = fig.add_subplot(3, 2, 5)
ax.plot(freq, zmeanx_psd, color='k', lw=1, alpha=.8)
ax.fill_between(freq[peak_indexes], zmeanx_psd[peak_indexes], color='blue', alpha=.5)
ax.set_ylim(0, zmeanx_psd.max()*1.15)
ax.set_xlim(0, freq.max())
ax.set_xlabel(r'frequency $f$ [frame$^{-1}$]')
ax.set_ylabel(r'[px$^2$ / frame$^{-1}$]')

# linear plot, zoomed near peak
ax = fig.add_subplot(3, 2, 6)
ax.plot(freq, zmeanx_psd, color='k', lw=1, alpha=.8)
ax.axvline(f0_guess, color='blue', linestyle='--')
# ax.axvline(peakedges[0], lw=2, color='blue', linestyle='--')
# ax.axvline(peakedges[1], lw=2, color='blue', linestyle='--')
ax.plot(freq[peak_indexes], zmeanx_psd[peak_indexes], color='k', lw=3)
ax.fill_between(freq[peak_indexes], zmeanx_psd[peak_indexes], color='blue', alpha=.5)
ax.set_ylim(0, zmeanx_psd[peak_indexes].max()*1.15)
ax.set_xlim(max(0, f0_guess - 4 * (peak_edges[1] - peak_edges[0])), f0_guess + 4 * (peak_edges[1] - peak_edges[0]))
ax.set_xlabel(r'frequency $f$ [frame$^{-1}$]')
ax.set_ylabel(r'[px$^2$ / frame$^{-1}$]')


# <codecell>

# Find k
# assumption: Z isd k-periodic

Z_kfinding = Z - Z.mean(axis=1, keepdims=True)

zero_pad_factor = (1,8)
window='hann'
if float(acquisition.split('a')[-1]) < 185:
    window='blackman'

k, f = utility.fourier.dual2d(x, t, zero_pad_factor=zero_pad_factor)
Z_kfinding_pw = utility.fourier.psd2d(Z, x, t, window=window, zero_pad_factor=zero_pad_factor)

power_k = Z_kfinding_pw.mean(axis=0)

size_of_0_peak = utility.fourier.peak_contour1d(0, power_k, x=k, peak_depth_dB=20)[1]

# region_search = k > size_of_0_peak
region_search = k > 0.002

q0_guess = utility.find_global_max(k[region_search], power_k[region_search])

print(f'q_0 (guessed): {round(q0_guess, 5)} px-1 ')


# <codecell>

plt.figure(figsize=utility.figsize('simple'))
plt.plot(k, power_k)
plt.axvline(size_of_0_peak, ls='--')
plt.axvline(q0_guess)
plt.gca().set_xlim(-3.5*q0_guess, 3.5*q0_guess)
plt.gca().set_yscale('log')


# <codecell>

# fz guess
Z_fzfinding = Z

zero_pad_factor = (16,1)
window='hann'

k, f = utility.fourier.dual2d(x, t, zero_pad_factor=zero_pad_factor)
Z_fzfinding_pw = utility.fourier.psd2d(Z, x, t, window=window, zero_pad_factor=zero_pad_factor)

power_f = Z_fzfinding_pw.mean(axis=1)

region_search = np.abs(f) < f0_guess/2

fz_guess = utility.find_global_max(f[region_search], utility.savgol_filter(power_f, 7, 2)[region_search])

# print(f'q_0 (guessed): {round(q0_guess, 5)} px-1 ')


# <codecell>

plt.figure(figsize=utility.figsize('simple'))
plt.plot(f, power_f)
plt.plot(f[region_search], utility.savgol_filter(power_f, 9, 2)[region_search], color='k')
plt.axvline(fz_guess)
# plt.axvline(q0_guess)
plt.gca().set_xlim(-3.5*f0_guess, 3.5*f0_guess)

plt.axvspan(f[region_search].min(), f[region_search].max(), alpha=.1, color='k', zorder=.5)


# <markdowncell>

# ## ##  # F# i# n# d#  # $# F# $


# <codecell>

zero_pad_factor_a = (8,8)
window_a='flattop'
zero_pad_factor_p = None
window_p='hann'
zero_pad_factor_fq = (8,8)
window_fq='boxcar'

plotix=False
if plotix:
    zero_pad_factor_fq = (4,4)
    zero_pad_factor_a = (4,4)

# pas besoin d'être super précis non plus
zero_pad_factor_fq = (4,4)
zero_pad_factor_a = (4,4)

q_fq, f_fq = utility.fourier.dual2d(x, t, zero_pad_factor=zero_pad_factor_fq)
qq_fq, ff_fq = np.meshgrid(q_fq, f_fq)

Zpw_fq = utility.fourier.psd2d(Z, x, t, window=window_fq, zero_pad_factor=zero_pad_factor_fq)
Wpw_fq = utility.fourier.psd2d(W, x, t, window=window_fq, zero_pad_factor=zero_pad_factor_fq)

q_a, f_a = utility.fourier.dual2d(x, t, zero_pad_factor=zero_pad_factor_a)
qq_a, ff_a = np.meshgrid(q_a, f_a)

Zpw_a = utility.fourier.psd2d(Z, x, t, window=window_a, zero_pad_factor=zero_pad_factor_a)
Wpw_a = utility.fourier.psd2d(W, x, t, window=window_a, zero_pad_factor=zero_pad_factor_a)

q_p, f_p = utility.fourier.dual2d(x, t, zero_pad_factor=zero_pad_factor_p)
qq_p, ff_p = np.meshgrid(q_p, f_p)

Zft_p = utility.fourier.ft2d(Z, x, t, window=window_p, zero_pad_factor=zero_pad_factor_p)
Wft_p = utility.fourier.ft2d(W, x, t, window=window_p, zero_pad_factor=zero_pad_factor_p)



# <codecell>

logging.set_verbose('debug')


# <codecell>

### F
# find f and k
F_f_guess = f0_guess
F_q_guess = 0
print(f'(F_f, F_q) = ({round(F_f_guess, 5)}, {round(F_q_guess, 5)}) [guess]')
peak_depth_dB = 20

F_peakcnt_fq = utility.fourier.peak_contour2d(F_q_guess, F_f_guess, Zpw_fq, peak_depth_dB=peak_depth_dB, x=q_fq, y=f_fq,
                                              real_signal=False)
F_peakvcn_fq = utility.fourier.peak_vicinity2d(F_q_guess, F_f_guess, Zpw_fq, peak_depth_dB=peak_depth_dB, x=q_fq, y=f_fq,
                                            peak_contours=F_peakcnt_fq)

condition = ff_fq > 0
F_q = np.average(qq_fq[condition], weights = (F_peakvcn_fq * Zpw_fq)[condition])
F_f = np.average(ff_fq[condition], weights = (F_peakvcn_fq * Zpw_fq)[condition])
print(f'(F_f, F_q) = ({round(F_f, 5)}, {round(F_q, 5)}) [found]')

# logging.set_verbose('debug')
# F_peakpw_fq = utility.fourier.power_near_peak2d(F_q_guess, F_f_guess, Zpw_fq, peak_depth_dB=peak_depth_dB, x=q_fq, y=f_fq,
#                                                  peak_contours=F_peakcnt_fq, peak_vicinity=F_peakvcn_fq) * 2
# print(f'Z_a = {round(F_peakpw_fq, 5)} px^2 [guess]')

peak_depth_dB = 20

F_peakcnt_a = utility.fourier.peak_contour2d(F_q, F_f, Zpw_a, peak_depth_dB=peak_depth_dB, x=q_a, y=f_a,
                                             real_signal=True, peak_max_area=1e3)
F_peakvcn_a = utility.fourier.peak_vicinity2d(F_q, F_f, Zpw_a, peak_depth_dB=peak_depth_dB, x=q_a, y=f_a,
                                              peak_contours=F_peakcnt_a)

F_peakpw_a = utility.fourier.power_near_peak2d(F_q_guess, F_f_guess, Zpw_a, peak_depth_dB=peak_depth_dB, x=q_a, y=f_a,
                                               peak_contours=F_peakcnt_a, peak_vicinity=F_peakvcn_a)
F_a = np.sqrt(F_peakpw_a)*np.sqrt(2)
print(f'F (power) = {round(F_peakpw_a, 5)} px^2 [measured]')
print(f'F_a = {round(F_a, 5)} px [measured]')



F_p = np.angle(Zft_p)[np.argmin((f_p - F_f) ** 2)][np.argmin((q_p - F_q) ** 2)]
print(f'F_p = {round(F_p, 5)} rad [measured]')


# <codecell>

if plotix:
    fig, axes = plt.subplots(3, 2, sharex=True, sharey=True)
    imshow_kw = {'origin':'upper',
                 'interpolation':'nearest',
                 'aspect':'auto'}
    
    ax = axes[0,0]
    vmax, vmin = utility.log_amplitude_range(Zpw_fq.max(), range_db=range_db)
    im_zpw = ax.imshow(Zpw_fq, extent=utility.correct_extent(q_fq, f_fq), norm='log', vmax=vmax, vmin=vmin, cmap='viridis', **imshow_kw)
    ax.set_xlabel(r'$k$ [px$^{-1}$]')
    ax.set_ylabel(r'$f$ [frame$^{-1}$]')
    cb = plt.colorbar(im_zpw, ax=ax, label=r'$|\hat{z}|^2$ [px^2/(px-1.frame-1)]')
    utility.set_ticks_log_cb(cb, vmax, range_db=range_db)
    ax.axhline(F_f_guess, ls='--')
    ax.axvline(F_q_guess, ls='--')
    ax.axhline(F_f, ls='-')
    ax.axvline(F_q, ls='-')
    
    
    for contour in F_peakcnt_fq:
        utility.draw_multipolygon_edge(ax, contour , lw=3, color='w')
        utility.draw_multipolygon_edge(ax, contour, lw=1, color='k')
    
    ax.set_xlim(-1/100, 1/100)
    ax.set_ylim(-1/50, 1/50)
    
    
    ax = axes[0,1]
    im_zpw = ax.imshow(F_peakvcn_fq*Zpw_fq, extent=utility.correct_extent(q_fq, f_fq), cmap='viridis', **imshow_kw)
    ax.axhline(F_f, ls='-')
    ax.axvline(F_q, ls='-')
    
    ax = axes[1,0]
    vmax, vmin = utility.log_amplitude_range(Zpw_a.max(), range_db=range_db)
    im_zpw = ax.imshow(Zpw_a, extent=utility.correct_extent(q_a, f_a), norm='log', vmax=vmax, vmin=vmin, cmap='viridis', **imshow_kw)
    ax.set_xlabel(r'$k$ [px$^{-1}$]')
    ax.set_ylabel(r'$f$ [frame$^{-1}$]')
    cb = plt.colorbar(im_zpw, ax=ax, label=r'$|\hat{z}|^2$ [px^2/(px-1.frame-1)]')
    utility.set_ticks_log_cb(cb, vmax, range_db=range_db)
    ax.axhline(F_f, ls='-')
    ax.axvline(F_q, ls='-')
    ax.axhline(-F_f, ls='-')
    ax.axvline(-F_q, ls='-')
    
    
    for contour in F_peakcnt_a:
        utility.draw_multipolygon_edge(ax, contour , lw=3, color='w')
        utility.draw_multipolygon_edge(ax, contour, lw=1, color='k')
    
    ax.set_xlim(-1/100, 1/100)
    ax.set_ylim(-1/50, 1/50)
    
    
    ax = axes[1,1]
    im_zpw = ax.imshow(F_peakvcn_a*Zpw_a, extent=utility.correct_extent(q_a, f_a), cmap='viridis', **imshow_kw)
    ax.axhline(F_f, ls='-')
    ax.axvline(F_q, ls='-')
    
    
    
    
    ax = axes[2,0]
    vmax, vmin = np.pi, -np.pi
    im_zpw = ax.imshow(np.angle(Zft_p), extent=utility.correct_extent(q_p, f_p), vmin=vmin, vmax=vmax, cmap='twilight_shifted', **imshow_kw)
    ax.set_xlabel(r'$k$ [px$^{-1}$]')
    ax.set_ylabel(r'$f$ [frame$^{-1}$]')
    cb = plt.colorbar(im_zpw, ax=ax, label=r'angle')
    # utility.set_ticks_log_cb(cb, vmax, range_db=range_db)
    ax.axhline(F_f, ls='-')
    ax.axvline(F_q, ls='-')
    ax.axhline(-F_f, ls='-')
    ax.axvline(-F_q, ls='-')
    
    
    ax = axes[2,1]
    im_zpw = ax.imshow(np.angle(Zft_p), alpha=np.abs(Zft_p)/np.max(np.abs(Zft_p)), extent=utility.correct_extent(q_p, f_p), vmin=vmin, vmax=vmax, cmap='twilight_shifted', **imshow_kw)
    ax.axhline(F_f, ls='-')
    ax.axvline(F_q, ls='-')
    ax.axhline(-F_f, ls='-')
    ax.axvline(-F_q, ls='-')
    # plt.tight_layout()


# <codecell>

### Z
# find f and k
Z_f_guess = fz_guess
Z_q_guess = q0_guess
print(f'(Z_f, Z_q) = ({round(Z_f_guess, 5)}, {round(Z_q_guess, 5)}) [guess]')
peak_depth_dB = 20

Z_peakcnt_fq = utility.fourier.peak_contour2d(Z_q_guess, Z_f_guess, Zpw_fq, peak_depth_dB=peak_depth_dB, x=q_fq, y=f_fq,
                                              real_signal=False)
Z_peakvcn_fq = utility.fourier.peak_vicinity2d(Z_q_guess, Z_f_guess, Zpw_fq, peak_depth_dB=peak_depth_dB, x=q_fq, y=f_fq,
                                               peak_contours=Z_peakcnt_fq)

condition = qq_fq > 0
Z_q = np.average(qq_fq[condition], weights = (Z_peakvcn_fq * Zpw_fq)[condition])
Z_f = np.average(ff_fq[condition], weights = (Z_peakvcn_fq * Zpw_fq)[condition])
print(f'(Z_f, Z_q) = ({round(Z_f, 5)}, {round(Z_q, 5)}) [found]')

# logging.set_verbose('debug')
# Z_peakpw_fq = utility.fourier.power_near_peak2d(Z_q_guess, Z_f_guess, Zpw_fq, peak_depth_dB=peak_depth_dB, x=q_fq, y=f_fq,
#                                                  peak_contours=Z_peakcnt_fq, peak_vicinity=Z_peakvcn_fq) * 2
# print(f'Z_a = {round(Z_peakpw_fq, 5)} px^2 [guess]')

peak_depth_dB = 20

Z_peakcnt_a = utility.fourier.peak_contour2d(Z_q, Z_f, Zpw_a, peak_depth_dB=peak_depth_dB, x=q_a, y=f_a,
                                             real_signal=True, peak_max_area=1e3)
Z_peakvcn_a = utility.fourier.peak_vicinity2d(Z_q, Z_f, Zpw_a, peak_depth_dB=peak_depth_dB, x=q_a, y=f_a,
                                              peak_contours=Z_peakcnt_a)

Z_peakpw_a = utility.fourier.power_near_peak2d(Z_q_guess, Z_f_guess, Zpw_a, peak_depth_dB=peak_depth_dB, x=q_a, y=f_a,
                                               peak_contours=Z_peakcnt_a, peak_vicinity=Z_peakvcn_a)
Z_a = np.sqrt(Z_peakpw_a)*np.sqrt(2)
print(f'Z (power) = {round(Z_peakpw_a, 5)} px^2 [measured]')
print(f'Z_a = {round(Z_a, 5)} px [measured]')



Z_p = np.angle(Zft_p)[np.argmin((f_p - Z_f) ** 2)][np.argmin((q_p - Z_q) ** 2)]
print(f'Z_p = {round(Z_p, 5)} rad [measured]')


# <codecell>

if plotix:
    fig, axes = plt.subplots(3, 2, sharex=True, sharey=True)
    imshow_kw = {'origin':'upper',
                 'interpolation':'nearest',
                 'aspect':'auto'}
    
    ax = axes[0,0]
    vmax, vmin = utility.log_amplitude_range(Zpw_fq.max(), range_db=range_db)
    im_zpw = ax.imshow(Zpw_fq, extent=utility.correct_extent(q_fq, f_fq), norm='log', vmax=vmax, vmin=vmin, cmap='viridis', **imshow_kw)
    ax.set_xlabel(r'$k$ [px$^{-1}$]')
    ax.set_ylabel(r'$f$ [frame$^{-1}$]')
    cb = plt.colorbar(im_zpw, ax=ax, label=r'$|\hat{z}|^2$ [px^2/(px-1.frame-1)]')
    utility.set_ticks_log_cb(cb, vmax, range_db=range_db)
    ax.axhline(Z_f_guess, ls='--')
    ax.axvline(Z_q_guess, ls='--')
    ax.axhline(Z_f, ls='-')
    ax.axvline(Z_q, ls='-')
    
    
    for contour in Z_peakcnt_fq:
        utility.draw_multipolygon_edge(ax, contour , lw=3, color='w')
        utility.draw_multipolygon_edge(ax, contour, lw=1, color='k')
    
    ax.set_xlim(-1/100, 1/100)
    ax.set_ylim(-1/50, 1/50)
    
    
    ax = axes[0,1]
    im_zpw = ax.imshow(Z_peakvcn_fq*Zpw_fq, extent=utility.correct_extent(q_fq, f_fq), cmap='viridis', **imshow_kw)
    ax.axhline(Z_f, ls='-')
    ax.axvline(Z_q, ls='-')
    
    ax = axes[1,0]
    vmax, vmin = utility.log_amplitude_range(Zpw_a.max(), range_db=range_db)
    im_zpw = ax.imshow(Zpw_a, extent=utility.correct_extent(q_a, f_a), norm='log', vmax=vmax, vmin=vmin, cmap='viridis', **imshow_kw)
    ax.set_xlabel(r'$k$ [px$^{-1}$]')
    ax.set_ylabel(r'$f$ [frame$^{-1}$]')
    cb = plt.colorbar(im_zpw, ax=ax, label=r'$|\hat{z}|^2$ [px^2/(px-1.frame-1)]')
    utility.set_ticks_log_cb(cb, vmax, range_db=range_db)
    ax.axhline(Z_f, ls='-')
    ax.axvline(Z_q, ls='-')
    ax.axhline(-Z_f, ls='-')
    ax.axvline(-Z_q, ls='-')
    
    
    for contour in Z_peakcnt_a:
        utility.draw_multipolygon_edge(ax, contour , lw=3, color='w')
        utility.draw_multipolygon_edge(ax, contour, lw=1, color='k')
    
    ax.set_xlim(-1/100, 1/100)
    ax.set_ylim(-1/50, 1/50)
    
    
    ax = axes[1,1]
    im_zpw = ax.imshow(Z_peakvcn_a*Zpw_a, extent=utility.correct_extent(q_a, f_a), cmap='viridis', **imshow_kw)
    ax.axhline(Z_f, ls='-')
    ax.axvline(Z_q, ls='-')
    
    
    ax = axes[2,0]
    vmax, vmin = np.pi, -np.pi
    im_zpw = ax.imshow(np.angle(Zft_p), extent=utility.correct_extent(q_p, f_p), vmin=vmin, vmax=vmax, cmap='twilight_shifted', **imshow_kw)
    ax.set_xlabel(r'$k$ [px$^{-1}$]')
    ax.set_ylabel(r'$f$ [frame$^{-1}$]')
    cb = plt.colorbar(im_zpw, ax=ax, label=r'angle')
    # utility.set_ticks_log_cb(cb, vmax, range_db=range_db)
    ax.axhline(Z_f, ls='-')
    ax.axvline(Z_q, ls='-')
    ax.axhline(-Z_f, ls='-')
    ax.axvline(-Z_q, ls='-')
    
    
    ax = axes[2,1]
    im_zpw = ax.imshow(np.angle(Zft_p), alpha=np.abs(Zft_p)/np.max(np.abs(Zft_p)), extent=utility.correct_extent(q_p, f_p), vmin=vmin, vmax=vmax, cmap='twilight_shifted', **imshow_kw)
    ax.axhline(Z_f, ls='-')
    ax.axvline(Z_q, ls='-')
    ax.axhline(-Z_f, ls='-')
    ax.axvline(-Z_q, ls='-')
    # plt.tight_layout()


# <codecell>

### W
# find f and k
W_f_guess = fz_guess + f0_guess
W_q_guess = q0_guess
print(f'(W_f, W_q) = ({round(W_f_guess, 5)}, {round(W_q_guess, 5)}) [guess]')
peak_depth_dB = 20

W_peakcnt_fq = utility.fourier.peak_contour2d(W_q_guess, W_f_guess, Wpw_fq, peak_depth_dB=peak_depth_dB, x=q_fq, y=f_fq,
                                              real_signal=False)
W_peakvcn_fq = utility.fourier.peak_vicinity2d(W_q_guess, W_f_guess, Wpw_fq, peak_depth_dB=peak_depth_dB, x=q_fq, y=f_fq,
                                               peak_contours=W_peakcnt_fq)

condition = qq_fq > 0
W_q = np.average(qq_fq[condition], weights = (W_peakvcn_fq * Wpw_fq)[condition])
W_f = np.average(ff_fq[condition], weights = (W_peakvcn_fq * Wpw_fq)[condition])
print(f'(W_f, W_q) = ({round(W_f, 5)}, {round(W_q, 5)}) [found]')

# logging.set_verbose('debug')
# W_peakpw_fq = utility.fourier.power_near_peak2d(W_q_guess, W_f_guess, Wpw_fq, peak_depth_dB=peak_depth_dB, x=q_fq, y=f_fq,
#                                                  peak_contours=W_peakcnt_fq, peak_vicinity=W_peakvcn_fq) * 2
# print(f'W_a = {round(W_peakpw_fq, 5)} px^2 [guess]')

peak_depth_dB = 20

W_peakcnt_a = utility.fourier.peak_contour2d(W_q, W_f, Wpw_a, peak_depth_dB=peak_depth_dB, x=q_a, y=f_a,
                                             real_signal=True, peak_max_area=1e3)
W_peakvcn_a = utility.fourier.peak_vicinity2d(W_q, W_f, Wpw_a, peak_depth_dB=peak_depth_dB, x=q_a, y=f_a,
                                              peak_contours=W_peakcnt_a)

W_peakpw_a = utility.fourier.power_near_peak2d(W_q_guess, W_f_guess, Wpw_a, peak_depth_dB=peak_depth_dB, x=q_a, y=f_a,
                                               peak_contours=W_peakcnt_a, peak_vicinity=W_peakvcn_a)
W_a = np.sqrt(W_peakpw_a)*np.sqrt(2)
print(f'W (power) = {round(W_peakpw_a, 5)} px^2 [measured]')
print(f'W_a = {round(W_a, 5)} px [measured]')



W_p = np.angle(Wft_p)[np.argmin((f_p - W_f) ** 2)][np.argmin((q_p - W_q) ** 2)]
print(f'W_p = {round(W_p, 5)} rad [measured]')


# <codecell>

if plotix:
    imshow_kw = {'origin':'upper',
                 'interpolation':'nearest',
                 'aspect':'auto'}
    fig, axes = plt.subplots(3, 2, sharex=True, sharey=True)
    
    ax = axes[0,0]
    vmax, vmin = utility.log_amplitude_range(Wpw_fq.max(), range_db=range_db)
    im_zpw = ax.imshow(Wpw_fq, extent=utility.correct_extent(q_fq, f_fq), norm='log', vmax=vmax, vmin=vmin, cmap='viridis', **imshow_kw)
    ax.set_xlabel(r'$k$ [px$^{-1}$]')
    ax.set_ylabel(r'$f$ [frame$^{-1}$]')
    cb = plt.colorbar(im_zpw, ax=ax, label=r'$|\hat{z}|^2$ [px^2/(px-1.frame-1)]')
    utility.set_ticks_log_cb(cb, vmax, range_db=range_db)
    ax.axhline(W_f_guess, ls='--')
    ax.axvline(W_q_guess, ls='--')
    ax.axhline(W_f, ls='-')
    ax.axvline(W_q, ls='-')
    
    
    for contour in W_peakcnt_fq:
        utility.draw_multipolygon_edge(ax, contour , lw=3, color='w')
        utility.draw_multipolygon_edge(ax, contour, lw=1, color='k')
    
    ax.set_xlim(-1/100, 1/100)
    ax.set_ylim(-1/50, 1/50)
    
    
    ax = axes[0,1]
    im_zpw = ax.imshow(W_peakvcn_fq*Wpw_fq, extent=utility.correct_extent(q_fq, f_fq), cmap='viridis', **imshow_kw)
    ax.axhline(W_f, ls='-')
    ax.axvline(W_q, ls='-')
    
    ax = axes[1,0]
    vmax, vmin = utility.log_amplitude_range(Wpw_a.max(), range_db=range_db)
    im_zpw = ax.imshow(Wpw_a, extent=utility.correct_extent(q_a, f_a), norm='log', vmax=vmax, vmin=vmin, cmap='viridis', **imshow_kw)
    ax.set_xlabel(r'$k$ [px$^{-1}$]')
    ax.set_ylabel(r'$f$ [frame$^{-1}$]')
    cb = plt.colorbar(im_zpw, ax=ax, label=r'$|\hat{z}|^2$ [px^2/(px-1.frame-1)]')
    utility.set_ticks_log_cb(cb, vmax, range_db=range_db)
    ax.axhline(W_f, ls='-')
    ax.axvline(W_q, ls='-')
    ax.axhline(-W_f, ls='-')
    ax.axvline(-W_q, ls='-')
    
    
    for contour in W_peakcnt_a:
        utility.draw_multipolygon_edge(ax, contour , lw=3, color='w')
        utility.draw_multipolygon_edge(ax, contour, lw=1, color='k')
    
    ax.set_xlim(-1/100, 1/100)
    ax.set_ylim(-1/50, 1/50)
    
    
    ax = axes[1,1]
    im_zpw = ax.imshow(W_peakvcn_a*Wpw_a, extent=utility.correct_extent(q_a, f_a), cmap='viridis', **imshow_kw)
    ax.axhline(W_f, ls='-')
    ax.axvline(W_q, ls='-')
    
    
    
    
    ax = axes[2,0]
    vmax, vmin = np.pi, -np.pi
    im_zpw = ax.imshow(np.angle(Wft_p), extent=utility.correct_extent(q_p, f_p), vmin=vmin, vmax=vmax, cmap='twilight_shifted', **imshow_kw)
    ax.set_xlabel(r'$k$ [px$^{-1}$]')
    ax.set_ylabel(r'$f$ [frame$^{-1}$]')
    cb = plt.colorbar(im_zpw, ax=ax, label=r'angle')
    # utility.set_ticks_log_cb(cb, vmax, range_db=range_db)
    ax.axhline(W_f, ls='-')
    ax.axvline(W_q, ls='-')
    ax.axhline(-W_f, ls='-')
    ax.axvline(-W_q, ls='-')
    
    
    ax = axes[2,1]
    im_zpw = ax.imshow(np.angle(Wft_p), alpha=np.abs(Wft_p)/np.max(np.abs(Wft_p)), extent=utility.correct_extent(q_p, f_p), vmin=vmin, vmax=vmax, cmap='twilight_shifted', **imshow_kw)
    ax.axhline(W_f, ls='-')
    ax.axvline(W_q, ls='-')
    ax.axhline(-W_f, ls='-')
    ax.axvline(-W_q, ls='-')
    # plt.tight_layout()


# <codecell>

plt.close()


# <codecell>




# <codecell>

### meta-info reading
import os
import pandas as pd

metainfo = pd.read_excel(os.path.join(dataset_path, 'datasheet.xlsx'), sheet_name='metainfo', skiprows=2)

meaningful_keys = [key for key in metainfo.keys() if 'unnamed' not in key.lower()]
valid = metainfo['acquisition_title'].astype(str) != 'nan'

acquisition_title = metainfo['acquisition_title'][valid].to_numpy()


# <codecell>

import pandas as pd
### AUTODATA WORKSHEET CREATION
sheet_name = 'autodata2'

workbookpath = os.path.join(dataset_path, 'datasheet.xlsx')

from openpyxl import load_workbook
 
### Check if sheet already exists
wb = load_workbook(workbookpath, read_only=True)   # open an Excel file and return a workbook
must_create_sheet = sheet_name not in wb.sheetnames
wb.close()

### Create sheet if we must.
if must_create_sheet:
    dataauto = {'acquisition_title': acquisition_title}
    # f: frequency [Hz]
    # q: spatial fdrequency [mm-1]
    # a: amplitude [mm]
    # p: phase [rad]
    
    # Forcing
    dataauto[f'F_f'] = np.full(len(acquisition_title), np.nan)
    dataauto[f'F_q'] = np.full(len(acquisition_title), np.nan)
    dataauto[f'F_a'] = np.full(len(acquisition_title), np.nan)
    dataauto[f'F_p'] = np.full(len(acquisition_title), np.nan)
    # Z transverse wave
    dataauto[f'Z_f'] = np.full(len(acquisition_title), np.nan)
    dataauto[f'Z_q'] = np.full(len(acquisition_title), np.nan)
    dataauto[f'Z_a'] = np.full(len(acquisition_title), np.nan)
    dataauto[f'Z_p'] = np.full(len(acquisition_title), np.nan)
    # W longitudinal wave
    dataauto[f'W_f'] = np.full(len(acquisition_title), np.nan)
    dataauto[f'W_q'] = np.full(len(acquisition_title), np.nan)
    dataauto[f'W_a'] = np.full(len(acquisition_title), np.nan)
    dataauto[f'W_p'] = np.full(len(acquisition_title), np.nan)
    
    df_autodata = pd.DataFrame(dataauto)
    
    with pd.ExcelWriter(workbookpath, engine='openpyxl', mode='a', if_sheet_exists='error') as writer:
        df_autodata.to_excel(writer, sheet_name=sheet_name, index=False)
        
### load sheet
df_autodata = pd.read_excel(workbookpath, sheet_name=sheet_name)

### find the right index corresponding to the current acquisition
if acquisition.replace('_gcv', '') in acquisition_title:
    i_acquisition = np.where(acquisition_title == acquisition.replace('_gcv', ''))[0][0]
else:
    raise('wtf the video is not in the meta data file ?')


# <codecell>


df_autodata.loc[i_acquisition, f'F_f'] = F_f
df_autodata.loc[i_acquisition, f'F_q'] = F_q
df_autodata.loc[i_acquisition, f'F_a'] = F_a
df_autodata.loc[i_acquisition, f'F_p'] = F_p
df_autodata.loc[i_acquisition, f'Z_f'] = Z_f
df_autodata.loc[i_acquisition, f'Z_q'] = Z_q
df_autodata.loc[i_acquisition, f'Z_a'] = Z_a
df_autodata.loc[i_acquisition, f'Z_p'] = Z_p
df_autodata.loc[i_acquisition, f'W_f'] = W_f
df_autodata.loc[i_acquisition, f'W_q'] = W_q
df_autodata.loc[i_acquisition, f'W_a'] = W_a
df_autodata.loc[i_acquisition, f'W_p'] = W_p


# <codecell>

with pd.ExcelWriter(workbookpath, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    df_autodata.to_excel(writer, sheet_name=sheet_name, index=False)


# <codecell>



