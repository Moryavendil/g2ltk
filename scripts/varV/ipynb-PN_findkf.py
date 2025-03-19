# -*- coding: utf-8 -*-
# <nbformat>3.0</nbformat>
# <codecell>

# %matplotlib notebook

import os
import numpy as np
import matplotlib.pyplot as plt

from g2ltk import set_verbose, datareading, datasaving, utility
utility.configure_mpl()


# <codecell>

### Datasets display
root_path = '../'
datasets = datareading.find_available_datasets(root_path)
print('Available datasets:', datareading.find_available_datasets(root_path))


# <codecell>

### Dataset selection & acquisitions display
dataset = '-'
if len(datasets) == 1:
    dataset = datasets[0]
    datareading.log_info(f'Auto-selected dataset {dataset}')
dataset_path = os.path.join(root_path, dataset)
datareading.describe_dataset(dataset_path, type='gcv', makeitshort=True)


# <codecell>

### Acquisition selection
acquisition = 'a200'
acquisition_path = os.path.join(dataset_path, acquisition)


# <codecell>

### Parameters definition

# conversion factor
px_per_mm = 0.
px_per_um = px_per_mm * 1e3

# parameters to find the rivulet
rivfinding_params = {
    'resize_factor': 2,
    'remove_median_bckgnd_zwise': True,
    'gaussian_blur_kernel_size': (1, 1), # (sigma_z, sigma_x)
    'white_tolerance': 70,
    'borders_min_distance': 5.,
    'borders_width': 6.,
    'max_rivulet_width': 150.,
}

# portion of the video that is of interest to us
framenumbers = np.arange(datareading.get_number_of_available_frames(acquisition_path))
roi = None, None, None, None  #start_x, start_y, end_x, end_y
if dataset=='40evo' and acquisition=='a350':
    framenumbers = np.arange(450)



# <codecell>

# Data fetching
datareading.describe_acquisition(dataset, acquisition, framenumbers = framenumbers, subregion=roi)

length, height, width = datareading.get_geometry(acquisition_path, framenumbers = framenumbers, subregion=roi)

t = datareading.get_t_frames(acquisition_path, framenumbers=framenumbers)
x = datareading.get_x_px(acquisition_path, framenumbers = framenumbers, subregion=roi, resize_factor=rivfinding_params['resize_factor'])

z_raw = datasaving.fetch_or_generate_data('bol', dataset, acquisition, framenumbers=framenumbers, roi=roi, **rivfinding_params)
w_raw = datasaving.fetch_or_generate_data('fwhmol', dataset, acquisition, framenumbers=framenumbers, roi=roi, **rivfinding_params)



# <codecell>

z_tmp = z_raw.copy()
w_tmp = w_raw.copy()


# <codecell>

from scipy.ndimage import gaussian_filter
blur_t_frame = 2 # blur in time (frame).
sigma_t = blur_t_frame
blur_x_px = 5 # blur in space (px).
sigma_x = blur_x_px


z_filtered = gaussian_filter(z_tmp, sigma=(sigma_t, sigma_x))
w_filtered = gaussian_filter(w_tmp, sigma=(sigma_t, sigma_x))

fig, axes = plt.subplots(2, 2, sharex=True, sharey=True, squeeze=False)
imshow_kw = {'aspect': 'auto', 'origin': 'upper'}

ax = axes[0, 0]
ax.set_title('Z (normal)')
imz = ax.imshow(z_tmp, extent=utility.correct_extent(x, t), cmap='viridis', **imshow_kw)
ax.set_xlabel('$x$ [px]')
ax.set_ylabel('$t$ [frame]')
plt.colorbar(imz, ax=ax, label='$z$ [px]')

ax = axes[0, 1]
ax.set_title('Z (smoothed)')
imz = ax.imshow(z_filtered, extent=utility.correct_extent(x, t), cmap='viridis', **imshow_kw)
ax.set_xlabel('$x$ [px]')
ax.set_ylabel('$t$ [frame]')
plt.colorbar(imz, ax=ax, label='$z$ [px]')

ax = axes[1, 0]
ax.set_title('W (normal)')
imz = ax.imshow(w_tmp, extent=utility.correct_extent(x, t), cmap='viridis', **imshow_kw)
ax.set_xlabel('$x$ [px]')
ax.set_ylabel('$t$ [frame]')
plt.colorbar(imz, ax=ax, label='$z$ [px]')

ax = axes[1, 1]
ax.set_title('W (smoothed)')
imz = ax.imshow(w_filtered, extent=utility.correct_extent(x, t), cmap='viridis', **imshow_kw)
ax.set_xlabel('$x$ [px]')
ax.set_ylabel('$t$ [frame]')
plt.colorbar(imz, ax=ax, label='$z$ [px]')

plt.tight_layout()

apply_gaussianfiler = True
if apply_gaussianfiler:
    z_tmp = z_filtered.copy()
    w_tmp = w_filtered.copy()
    utility.log_info('Gaussian filtering correction made')
else:
    utility.log_info('No gaussian filtering correction made')


# <codecell>

# Spatial cleaning : parabolic fit on the data
from scipy.signal import savgol_filter

# get temporal mean
z_xprofile = np.mean(z_tmp, axis=0)

# linear fit
rivs_fit_spatial = np.polyfit(x, z_xprofile, deg = 1)
utility.log_info(f'Position spatial drift linear estimation: {round(np.poly1d(rivs_fit_spatial)(x).max() - np.poly1d(rivs_fit_spatial)(x).min(),2)} px')

# savgol
savgol_width = len(x)//10 + (1 if len(x)%2 == 0 else 0)

"""
xcorrect can be 3 things
 * None: we do not want correction : it is biaised, or we are interested in the mean-time features
 * 'linear': we do a linear fit and remove everything from it
 * 'smoothed': we do a smoothing avec the average and remove that
 * 'total': total removal of the mean value
"""
xcorrect = 'linear'
if xcorrect is None:
    utility.log_info('No spatial correction made')
elif xcorrect == 'linear':
    utility.log_info('Linear spatial correction made')
    z_tmp = z_tmp - np.expand_dims(np.poly1d(rivs_fit_spatial)(x), axis=0)
elif xcorrect == 'smoothed':
    utility.log_info(f'Smoothed spatial correction made (savgol-{savgol_width}-2)')
    z_tmp = z_tmp - savgol_filter(z_xprofile, savgol_width, 2)
elif xcorrect == 'total':
    utility.log_info('Total spatial correction made')
    z_tmp = z_tmp - z_xprofile
else:
    utility.log_warning(f'What do you mean by xcorrect={xcorrect} ?')
    utility.log_info('No spatial correction made')


# plot
fig, axes = plt.subplots(2, 1, sharex=True, figsize=utility.figsize('double'))
ax = axes[0]
ax.plot(x, z_xprofile, color='k', alpha=0.5, label='old time-averaged riv position')
ax.plot(x, np.poly1d(rivs_fit_spatial)(x), color='r', alpha=0.5, label=f'linear fit')
ax.plot(x, savgol_filter(z_xprofile, savgol_width, 2), color='b', alpha=0.5, label=f'smooth')
ax.set_ylabel('z (px)')
ax.legend()

ax = axes[1]
ax.plot(x, z_tmp.mean(axis=0), color='k', label='New time-averaged riv position')
ax.set_xlabel('x (px)')
ax.set_ylabel('z (px)')
ax.legend()
plt.tight_layout()


# <codecell>

# Temporal cleaning : parabolic fit on the data

# get spatial mean
z_tprofile = np.mean(z_tmp, axis=1)

# fit
rivs_fit_temporal = np.polyfit(t, z_tprofile, deg = 2)

# correct
do_tcorrect= False
if do_tcorrect:
    z_tmp = z_tmp - np.expand_dims(np.poly1d(rivs_fit_temporal)(t), axis=1)
else:
    utility.log_info('No temporal correction made')

# plot
plt.figure(figsize=(8,3))
ax = plt.gca()
if do_tcorrect:
    ax.plot(t, z_tprofile, color='k', alpha=0.5, label='old space-averaged riv position')
    plt.plot(t, np.poly1d(rivs_fit_temporal)(t), color='r', alpha=0.5, label=f'paraboloidal fit')
ax.plot(t, z_tmp.mean(axis=1), color='k', label='space-averaged riv position')
ax.set_xlabel('t (s)')
ax.set_ylabel('z (px)')
ax.legend()
plt.tight_layout()


# <codecell>

Z = z_tmp.copy()
W = w_tmp.copy()

zero_pad_factor = (5,5)
window='hann'

k, f = utility.fourier.dual2d(x, t, zero_pad_factor=zero_pad_factor)
Z_pw = utility.fourier.psd2d(Z, x, t, window=window, zero_pad_factor=zero_pad_factor)
W_pw = utility.fourier.psd2d(W, x, t, window=window, zero_pad_factor=zero_pad_factor)

range_db = 100


# <codecell>

# fig, axes = plt.subplots(2, 2)
# imshow_kw = {'origin':'upper', 
#              'interpolation':'nearest', 
#              'aspect':'auto'}
# 
# ax = axes[0,0]
# ax.set_title('Z (normal)')
# imz = ax.imshow(Z, extent=utility.correct_extent(x, t), cmap='viridis', **imshow_kw)
# ax.set_xlabel('$x$ [px]')
# ax.set_ylabel('$t$ [frame]')
# plt.colorbar(imz, ax=ax, label='$z$ [px]')
# 
# ax = axes[0,1]
# vmax, vmin = utility.log_amplitude_range(Z_pw.max(), range_db=range_db)
# im_zpw = ax.imshow(Z_pw, extent=utility.correct_extent(k, f), norm='log', vmax=vmax, vmin=vmin, cmap='viridis', **imshow_kw)
# ax.set_xlabel(r'$k$ [px$^{-1}$]')
# ax.set_ylabel(r'$f$ [frame$^{-1}$]')
# cb = plt.colorbar(im_zpw, ax=ax, label=r'$|\hat{z}|^2$ [px^2/(px-1.frame-1)]')
# utility.set_ticks_log_cb(cb, vmax, range_db=range_db)
# 
# ax.set_xlim(0, 1/50)
# ax.set_ylim(-1/20, 1/20)
# 
# ax = axes[1,0]
# ax.set_title('W (normal)')
# imz = ax.imshow(W, extent=utility.correct_extent(x, t), cmap='viridis', **imshow_kw)
# ax.set_xlabel('$x$ [px]')
# ax.set_ylabel('$t$ [frame]')
# plt.colorbar(imz, ax=ax, label='$w$ [px]')
# 
# ax = axes[1,1]
# vmax, vmin = utility.log_amplitude_range(W_pw.max(), range_db=range_db)
# im_zpw = ax.imshow(W_pw, extent=utility.correct_extent(k, f), norm='log', vmax=vmax, vmin=vmin, cmap='viridis', **imshow_kw)
# ax.set_xlabel(r'$k$ [px$^{-1}$]')
# ax.set_ylabel(r'$f$ [frame$^{-1}$]')
# cb = plt.colorbar(im_zpw, ax=ax, label=r'$|\hat{w}|^2$ [px^2/(px-1.frame-1)]')
# utility.set_ticks_log_cb(cb, vmax, range_db=range_db)
# 
# ax.set_xlim(0, 1/50)
# ax.set_ylim(-1/20, 1/20)
# 
# # plt.tight_layout()


# <markdowncell>

# ## Find $f$, $k$, $\nu$
# 
# Find the parameters of the PN instability


# <markdowncell>

# ## ##  # F# i# n# d#  # $# f# _# 0# $


# <codecell>

### HERE WE FIND THE EXCITATION FREQUENCY BY CONSIDERING THE SPACE-AVERAGED VERSION OF Z
window = 'hann'
peak_depth_dB = 60

# Take the space-average
zmeanx = np.mean(Z, axis=1)

# Compute the power spectral density
freq = utility.rdual(t, zero_pad_factor=zero_pad_factor[0])
zmeanx_psd = utility.psd1d(zmeanx, t, window=window, zero_pad_factor=zero_pad_factor[0])  # power spectral density

# find the main peak
f0_guess = utility.fourier.estimatesignalfrequency(zmeanx, t, window='boxcar')
print(f'f_0 (guessed): {round(f0_guess, 3)} frames-1 ')

peak_edges = utility.peak_contour1d(f0_guess, zmeanx_psd, peak_depth_dB=peak_depth_dB, x=freq)
peak_indexes = utility.peak_vicinity1d(f0_guess, zmeanx_psd, peak_depth_dB=peak_depth_dB, x=freq)
z_peakpower = utility.power_near_peak1d(f0_guess, zmeanx_psd, peak_depth_dB=peak_depth_dB, x=freq)

z_peakamplitude = np.sqrt(z_peakpower) * np.sqrt(2)
print(f'z_0: {round(z_peakamplitude, 3)} px (filtering PSD of <Z>_x)')
z0_measure = z_peakamplitude


# <codecell>

from scipy.signal import hilbert

# analytic_signal = hilbert(zmeanx) # this is brutal (but sometimes works well. it avoid the phase bad definition if at the end we are not on a min / max
analytic_signal = hilbert(np.concatenate((zmeanx, zmeanx[::-1])))[:len(zmeanx)] # We use a small symmetrization trick here bcz hilbert thinks everything in life has to be periodic smh
h = np.abs(analytic_signal)

fig = plt.figure()

# real signal 
ax = fig.add_subplot(3, 1, 1)
ax.plot(t, zmeanx, color='k')
ax.axhline(z_peakamplitude, lw=2, color='b', linestyle='--', label='$z_0$')
ax.axhline(-z_peakamplitude, lw=2, color='b', linestyle='--')
# ax.plot(t, h, color='m', alpha=.5, label='Analytic signal amplitude (hilbert)')
# ax.plot(t, -h, color='m', alpha=.5)
ax.set_xlabel(r'time $t$ [frame]')
ax.set_ylabel(r'<z>$_x$ [px]')
ax.legend()

# log plot, global
ax = fig.add_subplot(3, 2, 3)
ax.plot(freq, zmeanx_psd, color='k', lw=1, alpha=.8)
ax.fill_between(freq[peak_indexes], zmeanx_psd[peak_indexes], color='blue', alpha=.5)
ax.set_yscale('log')
ax.set_ylim(utility.attenuate_power(zmeanx_psd.max(), 200), zmeanx_psd.max()*2)
ax.set_xlim(0, freq.max())
ax.set_xlabel(r'frequency $f$ [frame$^{-1}$]')
ax.set_ylabel(r'[px$^2$ / frame$^{-1}$]')

# log plot, zoomed near peak
ax = fig.add_subplot(3, 2, 4)
ax.plot(freq, zmeanx_psd, color='k', lw=1, alpha=.8)
ax.axvline(f0_guess, color='blue', linestyle='--', label='Forcing freq')
# ax.axvline(peakedges[0], lw=2, color='blue', linestyle='--')
# ax.axvline(peakedges[1], lw=2, color='blue', linestyle='--')
ax.plot(freq[peak_indexes], zmeanx_psd[peak_indexes], color='k', lw=3)
ax.fill_between(freq[peak_indexes], zmeanx_psd[peak_indexes], color='blue', alpha=.5)
ax.set_yscale('log')
ax.set_ylim(utility.attenuate_power(zmeanx_psd[peak_indexes].max(), range_db+40), zmeanx_psd[peak_indexes].max()*2)
ax.set_xlim(max(0, f0_guess - 4 * (peak_edges[1] - peak_edges[0])), f0_guess + 4 * (peak_edges[1] - peak_edges[0]))
ax.set_xlabel(r'frequency $f$ [frame$^{-1}$]')
ax.set_ylabel(r'[px$^2$ / frame$^{-1}$]')
ax.legend()

# linear plot, gloabal
ax = fig.add_subplot(3, 2, 5)
ax.plot(freq, zmeanx_psd, color='k', lw=1, alpha=.8)
ax.fill_between(freq[peak_indexes], zmeanx_psd[peak_indexes], color='blue', alpha=.5)
ax.set_ylim(0, zmeanx_psd.max()*1.15)
ax.set_xlim(0, freq.max())
ax.set_xlabel(r'frequency $f$ [frame$^{-1}$]')
ax.set_ylabel(r'[px$^2$ / frame$^{-1}$]')

# linear plot, zoomed near peak
ax = fig.add_subplot(3, 2, 6)
ax.plot(freq, zmeanx_psd, color='k', lw=1, alpha=.8)
ax.axvline(f0_guess, color='blue', linestyle='--')
# ax.axvline(peakedges[0], lw=2, color='blue', linestyle='--')
# ax.axvline(peakedges[1], lw=2, color='blue', linestyle='--')
ax.plot(freq[peak_indexes], zmeanx_psd[peak_indexes], color='k', lw=3)
ax.fill_between(freq[peak_indexes], zmeanx_psd[peak_indexes], color='blue', alpha=.5)
ax.set_ylim(0, zmeanx_psd[peak_indexes].max()*1.15)
ax.set_xlim(max(0, f0_guess - 4 * (peak_edges[1] - peak_edges[0])), f0_guess + 4 * (peak_edges[1] - peak_edges[0]))
ax.set_xlabel(r'frequency $f$ [frame$^{-1}$]')
ax.set_ylabel(r'[px$^2$ / frame$^{-1}$]')


# <markdowncell>

# ## ## ##  # F# i# n# d#  # $# k# $


# <codecell>

### HERE WE FIND THE WAVENUMBER BY LOOKING AT THE AVERAGED SPECTRUM OF Z
window = 'hann'
peak_depth_dB = 60

# Take the time-average
zmeant = np.mean(Z, axis=0)

# Compute the power spectral density
freqx = utility.rdual(x, zero_pad_factor=zero_pad_factor[1])
zmeant_psd = utility.psd1d(zmeant, x, window=window, zero_pad_factor=zero_pad_factor[1])  # power spectral density

# find the main peak
k0_guess = utility.fourier.estimatesignalfrequency(zmeant, x)
if dataset=='40evo' and acquisition in ['a200', 'a190', 'a180']:
    k0_guess = utility.fourier.estimatesignalfrequency(zmeant, x, window='hann', bounds = [0.00225, 0.00325])
    peak_depth_dB = 40
if dataset=='40evo' and acquisition in ['a170', 'a160', 'a150', 'a140', 'a130', 'a120']:
    k0_guess = utility.fourier.estimatesignalfrequency(zmeant, x, window='hann', bounds = [0.00225, 0.00325])
    peak_depth_dB = 20
print(f'k_0 (guessed): {round(k0_guess, 3)} px-1 ')

peak_edges = utility.peak_contour1d(k0_guess, zmeant_psd, peak_depth_dB=peak_depth_dB, x=freqx)
peak_indexes = utility.peak_vicinity1d(k0_guess, zmeant_psd, peak_depth_dB=peak_depth_dB, x=freqx)
z_peakpower = utility.power_near_peak1d(k0_guess, zmeant_psd, peak_depth_dB=peak_depth_dB, x=freqx)
# 
# z_peakamplitude = np.sqrt(z_peakpower) * np.sqrt(2)
# print(f'z_0: {round(z_peakamplitude, 3)} px (filtering PSD of <Z>_x)')


# <codecell>

from scipy.signal import hilbert

# analytic_signal = hilbert(zmeanx) # this is brutal (but sometimes works well. it avoid the phase bad definition if at the end we are not on a min / max
analytic_signal = hilbert(np.concatenate((zmeanx, zmeanx[::-1])))[:len(zmeanx)] # We use a small symmetrization trick here bcz hilbert thinks everything in life has to be periodic smh
h = np.abs(analytic_signal)

fig = plt.figure()

# real signal 
ax = fig.add_subplot(3, 1, 1)
ax.plot(x, zmeant, color='k')
ax.axhline(z_peakamplitude, lw=2, color='b', linestyle='--', label='$z_0$')
ax.axhline(-z_peakamplitude, lw=2, color='b', linestyle='--')
# ax.plot(t, h, color='m', alpha=.5, label='Analytic signal amplitude (hilbert)')
# ax.plot(t, -h, color='m', alpha=.5)
ax.set_xlabel(r'space $x$ [px]')
ax.set_ylabel(r'<z>$_x$ [px]')
ax.legend()

# log plot, global
ax = fig.add_subplot(3, 2, 3)
ax.plot(freqx, zmeant_psd, color='k', lw=1, alpha=.8)
ax.fill_between(freqx[peak_indexes], zmeant_psd[peak_indexes], color='blue', alpha=.5)
ax.set_yscale('log')
ax.set_ylim(utility.attenuate_power(zmeant_psd.max(), 200), zmeant_psd.max()*2)
ax.set_xlim(0, min(freqx.max(), k0_guess*20))
ax.set_xlabel(r'frequency $f$ [frame$^{-1}$]')
ax.set_ylabel(r'[px$^2$ / frame$^{-1}$]')

# log plot, zoomed near peak
ax = fig.add_subplot(3, 2, 4)
ax.plot(freqx, zmeant_psd, color='k', lw=1, alpha=.8)
ax.axvline(k0_guess, color='blue', linestyle='--', label='k')
# ax.axvline(peakedges[0], lw=2, color='blue', linestyle='--')
# ax.axvline(peakedges[1], lw=2, color='blue', linestyle='--')
ax.plot(freqx[peak_indexes], zmeant_psd[peak_indexes], color='k', lw=3)
ax.fill_between(freqx[peak_indexes], zmeant_psd[peak_indexes], color='blue', alpha=.5)
ax.set_yscale('log')
ax.set_ylim(utility.attenuate_power(zmeant_psd[peak_indexes].max(), range_db+40), zmeant_psd[peak_indexes].max()*2)
ax.set_xlim(max(0, k0_guess - 3 * (peak_edges[1] - peak_edges[0])), k0_guess + 3 * (peak_edges[1] - peak_edges[0]))
ax.set_xlabel(r'$k$ [px$^{-1}$]')
ax.set_ylabel(r'[px$^2$ / frame$^{-1}$]')
ax.legend()

# linear plot, gloabal
ax = fig.add_subplot(3, 2, 5)
ax.plot(freqx, zmeant_psd, color='k', lw=1, alpha=.8)
ax.fill_between(freqx[peak_indexes], zmeant_psd[peak_indexes], color='blue', alpha=.5)
ax.set_ylim(0, zmeant_psd.max()*1.15)
ax.set_xlim(0, min(freqx.max(), k0_guess*20))
ax.set_xlabel(r'$k$ [px$^{-1}$]')
ax.set_ylabel(r'[px$^2$ / frame$^{-1}$]')

# linear plot, zoomed near peak
ax = fig.add_subplot(3, 2, 6)
ax.plot(freqx, zmeant_psd, color='k', lw=1, alpha=.8)
ax.axvline(k0_guess, color='blue', linestyle='--')
# ax.axvline(peakedges[0], lw=2, color='blue', linestyle='--')
# ax.axvline(peakedges[1], lw=2, color='blue', linestyle='--')
ax.plot(freqx[peak_indexes], zmeant_psd[peak_indexes], color='k', lw=3)
ax.fill_between(freqx[peak_indexes], zmeant_psd[peak_indexes], color='blue', alpha=.5)
ax.set_ylim(0, zmeant_psd[peak_indexes].max()*1.15)
ax.set_xlim(max(0, k0_guess - 3 * (peak_edges[1] - peak_edges[0])), k0_guess + 3 * (peak_edges[1] - peak_edges[0]))
ax.set_xlabel(r'$k$ [px$^{-1}$]')
ax.set_ylabel(r'[px$^2$ / frame$^{-1}$]')


# <markdowncell>

# ## ## ##  # F# i# n# d#  # $# \# n# u# $#  # (# v#  # d# r# i# f# t# )


# <codecell>

### HERE WE FIND VDRIFT

### HERE WE FIND THE EXCITATION FREQUENCY BY CONSIDERING THE SPACE-AVERAGED VERSION OF Z
k_nearest = np.argmin((k-k0_guess)**2)
k_indexes = [k_nearest + i for i in [-1, 0, 1]]

zpsd_k0 = np.mean(Z_pw[:, k_indexes], axis=1)

xlim = [-f0_guess, f0_guess]
in_xlim = (f > xlim[0]) * (f < xlim[1])

fdrift_guess = utility.find_global_peak(f[in_xlim], np.log10(zpsd_k0[in_xlim]), 'max')

vdrift = -fdrift_guess / k0_guess

print(f'fdrift (guessed): {round(fdrift_guess, 3)} frame-1 ')
print(f'vdrift (guessed): {round(k0_guess, 3)} px/frame ')


# <codecell>

fig = plt.figure()

ax = fig.add_subplot(211)
ax.plot(f, zpsd_k0, color='k')
for i_f0 in range(int(f.min()/f0_guess) - 1, int(f.max()/f0_guess) + 1 + 1):
    ax.axvline(fdrift_guess + i_f0*f0_guess, lw=1, color='b', linestyle='-', alpha=1.)
ax.set_yscale('log')
ax.set_xlim(-5*f0_guess, 5*f0_guess)
ax.set_xlabel(r'$f$ [frame$^{-1}$]')
# ax.set_ylabel(r'$f$-averaged PSD of Z [px^2/frame$^{-1}$]')
# ax.legend()

ax = fig.add_subplot(223)
ax.plot(f, zpsd_k0, color='k')
# ax.axvline(firstpeak_coarse, lw=1, color='b', linestyle='--', alpha=0.5, label='$k$ (coarse estimation)')
ax.axvline(fdrift_guess, lw=1, color='b', linestyle='-', alpha=1., label='$fdrift$ (fine estimation)')
ax.set_xlim(xlim)
ax.set_xlabel(r'$f$ [frame$^{-1}$]')
# ax.set_ylabel(r'$f$-averaged PSD of Z [px^2/frame$^{-1}$]')
# ax.legend()

ax = fig.add_subplot(224)
ax.plot(f, zpsd_k0, color='k')
# ax.axvline(firstpeak_coarse, lw=1, color='b', linestyle='--', alpha=0.5, label='$k$ (coarse estimation)')
ax.axvline(fdrift_guess, lw=1, color='b', linestyle='-', alpha=1., label='$fdrift$ (fine estimation)')
ax.set_yscale('log')
ax.set_xlim(xlim)
ax.set_xlabel(r'$f$ [frame$^{-1}$]')
# ax.set_ylabel(r'$f$-averaged PSD of Z [px^2/frame$^{-1}$]')
# ax.legend()


# <markdowncell>

# ## ##  # F# i# n# d#  # $# z# _# 1# $#  # a# n# d#  # $# w# _# 1# $


# <codecell>

# Z = z_tmp.copy()
# W = w_tmp.copy()
# 
# zero_pad_factor = (5,5)
# window='hann'
# 
# k, f = utility.fourier.dual2d(x, t, zero_pad_factor=zero_pad_factor)
# Z_pw = utility.fourier.psd2d(Z, x, t, window=window, zero_pad_factor=zero_pad_factor)
# W_pw = utility.fourier.psd2d(W, x, t, window=window, zero_pad_factor=zero_pad_factor)
# 
# range_db = 100


# <codecell>

kcentre_z = k0_guess
fcentre_z = fdrift_guess

kcentre_w = k0_guess
fcentre_w = f0_guess + fdrift_guess

peak_depth_dB = 40

options = {'peak_max_area': 100*zero_pad_factor[0]*zero_pad_factor[1], 'peak_min_circularity': .4}

z_peakpower = 0.
w_peakpower = 0.
if float(acquisition.split('a')[1].split('_')[0]) > 170:
    z_peakcontours = utility.peak_contour2d(kcentre_z, fcentre_z, Z_pw, peak_depth_dB, x=k, y=f, **options)
    z_peakmask = utility.peak_vicinity2d(kcentre_z, fcentre_z, Z_pw, peak_depth_dB, x=k, y=f, peak_contours=z_peakcontours, **options)
    z_peakpower = utility.power_near_peak2d(kcentre_z, fcentre_z, Z_pw, peak_depth_dB, x=k, y=f, peak_vicinity=z_peakmask, **options)
    
    w_peakcontours = utility.peak_contour2d(kcentre_w, fcentre_w, W_pw, peak_depth_dB, x=k, y=f, **options)
    w_peakmask = utility.peak_vicinity2d(kcentre_w, fcentre_w, W_pw, peak_depth_dB, x=k, y=f, peak_contours=w_peakcontours, **options)
    w_peakpower = utility.power_near_peak2d(kcentre_w, fcentre_w, W_pw, peak_depth_dB, x=k, y=f, peak_vicinity=w_peakmask, **options)

z_peakamplitude = np.sqrt(z_peakpower) * np.sqrt(2)
w_peakamplitude = np.sqrt(w_peakpower) * np.sqrt(2)
z1_measure, w1_measure = z_peakamplitude, w_peakamplitude
utility.log_info(f'z_1 = {round(z_peakamplitude, 3)} px (filtering PSD of Z)')
utility.log_info(f'w_1 = {round(w_peakamplitude, 3)} px (filtering PSD of W)')


# <codecell>

# fig, axes = plt.subplots(1, 2, squeeze=False, sharey=True, sharex=True)
# imshow_kw = {'origin':'upper', 
#              'interpolation':'nearest', 
#              'aspect':'auto'}
# fftplot_kw = {'origin': 'lower', 'aspect': 'auto', 'interpolation': 'nearest', 
#               'norm': 'log',
#               'extent': utility.correct_extent(k, f, origin='lower')}
# 
# 
# ax = axes[0,0]
# vmax, vmin = utility.log_amplitude_range(Z_pw.max(), range_db=range_db)
# im_zpw = ax.imshow(Z_pw, vmax=vmax, vmin=vmin, cmap='viridis', **fftplot_kw)
# ax.set_xlabel(r'$k$ [px$^{-1}$]')
# ax.set_ylabel(r'$f$ [frame$^{-1}$]')
# cb = plt.colorbar(im_zpw, ax=ax, label=r'$|\hat{z}|^2$ [px^2/(px-1.frame-1)]')
# utility.set_ticks_log_cb(cb, vmax, range_db=range_db)
# 
# ax.axvline(k0_guess, color='w', linestyle=':', alpha=.8)
# ax.axhline(fdrift_guess, color='w', linestyle=':', alpha=.8)
# for contour in z_peakcontours:
#     utility.draw_multipolygon_edge(ax, contour, xmin=-utility.step(k)/2, lw=3, color='w')
#     utility.draw_multipolygon_edge(ax, contour, xmin=-utility.step(k)/2, lw=1, color='k')
# 
# ax = axes[0,1]
# vmax, vmin = utility.log_amplitude_range(W_pw.max(), range_db=range_db)
# im_zpw = ax.imshow(W_pw, vmax=vmax, vmin=vmin, cmap='viridis', **fftplot_kw)
# ax.set_xlabel(r'$k$ [px$^{-1}$]')
# ax.set_ylabel(r'$f$ [frame$^{-1}$]')
# cb = plt.colorbar(im_zpw, ax=ax, label=r'$|\hat{w}|^2$ [px^2/(px-1.frame-1)]')
# utility.set_ticks_log_cb(cb, vmax, range_db=range_db)
# 
# ax.axvline(k0_guess, color='w', linestyle=':', alpha=.8)
# ax.axhline(f0_guess + fdrift_guess, color='w', linestyle=':', alpha=.8)
# for contour in w_peakcontours:
#     utility.draw_multipolygon_edge(ax, contour, xmin=-utility.step(k)/2, lw=3, color='w')
#     utility.draw_multipolygon_edge(ax, contour, xmin=-utility.step(k)/2, lw=1, color='k')
# 
# ax.set_xlim(0, 3*k0_guess)
# ax.set_ylim(-2*f0_guess + fdrift_guess, 3*f0_guess + fdrift_guess)
# 
# # plt.tight_layout()


# <markdowncell>

# ## Saving 
# 
# We are going to save now


# <codecell>

### meta-info reading
import pandas as pd

metainfo = pd.read_excel(os.path.join(dataset_path, 'datasheet.xlsx'), sheet_name='metainfo', skiprows=2)

meaningful_keys = [key for key in metainfo.keys() if 'unnamed' not in key.lower()]
valid = metainfo['acquisition_title'].astype(str) != 'nan'

acquisition_title = metainfo['acquisition_title'][valid].to_numpy()


# <codecell>

import pandas as pd
### AUTODATA WORKSHEET CREATION
sheet_name = 'autodata'

workbookpath = os.path.join(dataset_path, 'datasheet.xlsx')

from openpyxl import load_workbook
 
### Check if sheet already exists
wb = load_workbook(workbookpath, read_only=True)   # open an Excel file and return a workbook
must_create_sheet = sheet_name not in wb.sheetnames
wb.close()

### Create sheet if we must.
if must_create_sheet:
    dataauto = {'acquisition_title': acquisition_title}
    # fknu
    dataauto[f'f0'] = np.full(len(acquisition_title), np.nan)
    dataauto[f'q0'] = np.full(len(acquisition_title), np.nan)
    dataauto[f'fdrift'] = np.full(len(acquisition_title), np.nan)
    # amplitudes
    dataauto[f'z0'] = np.full(len(acquisition_title), np.nan)
    dataauto[f'z1'] = np.full(len(acquisition_title), np.nan)
    dataauto[f'w1'] = np.full(len(acquisition_title), np.nan)
    
    df_autodata = pd.DataFrame(dataauto)
    
    with pd.ExcelWriter(workbookpath, engine='openpyxl', mode='a', if_sheet_exists='error') as writer:
        df_autodata.to_excel(writer, sheet_name=sheet_name, index=False)
        
### load sheet
df_autodata = pd.read_excel(workbookpath, sheet_name=sheet_name)

### find the right index corresponding to the current acquisition
if acquisition.replace('_gcv', '') in acquisition_title:
    i_acquisition = np.where(acquisition_title == acquisition.replace('_gcv', ''))[0][0]
else:
    raise('wtf the video is not in the meta data file ?')


# <codecell>


df_autodata.loc[i_acquisition, f'f0'] = f0_guess
df_autodata.loc[i_acquisition, f'q0'] = k0_guess
df_autodata.loc[i_acquisition, f'fdrift'] = fdrift_guess
df_autodata.loc[i_acquisition, f'z0'] = z0_measure
df_autodata.loc[i_acquisition, f'z1'] = z1_measure
df_autodata.loc[i_acquisition, f'w1'] = w1_measure


# <codecell>

with pd.ExcelWriter(workbookpath, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    df_autodata.to_excel(writer, sheet_name=sheet_name, index=False)

